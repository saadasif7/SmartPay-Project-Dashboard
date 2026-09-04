import streamlit as st
import plotly.express as px
from datetime import datetime
import requests
from utils.excel_reader import load_data, get_weather
from utils.voice import listen
import streamlit.components.v1 as components
from rapidfuzz import process, fuzz
from io import BytesIO
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image
from utils.pdf_report import generate_executive_pdf




def speak(text):

    text = str(text).replace('"', '\\"')

    components.html(
        f"""
        <script>

        window.speechSynthesis.cancel();

        var msg = new SpeechSynthesisUtterance("{text}");

        msg.lang = "en-US";
        msg.rate = 0.95;
        msg.pitch = 1;
        msg.volume = 1;

        window.speechSynthesis.speak(msg);

        </script>
        """,
        height=0,
    )

# =====================================================
# Normalize Voice
# =====================================================

def normalize_voice(text):

    if not text:
        return ""

    text = str(text).lower().strip()

    aliases = {

        # ==========================
        # Staff
        # ==========================

        "sada asif": "saad asif",
        "sad asif": "saad asif",
        "saada asif": "saad asif",
        "saad asifa": "saad asif",
        "saad asif": "saad asif",

        "show all projects": "all",
        "show all": "all",
        "all projects": "all",

        "afzal": "afzal",

        "adnan": "adnan",
        "adnan younis": "adnan younis",

        "mobashir": "mubashir",
        "mubashar": "mubashir",
        "mubashir": "mubashir",

        "bhawan": "bhawan",
        "bawan": "bhawan",
        "bhavan": "bhawan",

        "ayesha": "ayesha",

        # ==========================
        # Status
        # ==========================

        "life": "live",
        "live": "live",

        "uat": "uat",
        "sit": "sit",

        "scoping": "under scoping",
        "under scoping": "under scoping",

        "development": "under development",
        "under development": "under development",

        "process": "in process",
        "in process": "in process",

        # ==========================
        # Projects
        # ==========================

        "swap": "swap",
        "swaps": "swap",

        "ata": "cnic screening -ata",

        "siem": "siem intergration",
        "cm": "siem intergration",

        "soap": "discontinuation of soap to rest",

        "token": "tokinization on smartpay",
        "tokenization": "tokinization on smartpay",

        "estamping": "e-stamping",
        "e stamping": "e-stamping",

        "outward": "outward clearing",

        "gbm": "gbm",

        "optimization": "optimization",

        "rf": "rf account on smartpay",
        "show live projects": "live",
        "live projects": "live",
        "only live": "live",

        "show uat projects": "uat",
        "uat projects": "uat",

        "show sit projects": "sit",
        "sit projects": "sit",

        "show development projects": "under development",
        "under development": "under development",

        "show all projects": "all",
        "all projects": "all",
         # =====================================
        # Voice Commands
        # =====================================

        "show all projects": "all",
        "show all": "all",
        "all projects": "all",
        "all": "all",

        "show live projects": "live",
        "live projects": "live",
        "only live": "live",

        "show uat projects": "uat",
        "uat projects": "uat",

        "show sit projects": "sit",
        "sit projects": "sit",

        "show development projects": "under development",
        "development projects": "under development",

        # ==========================
        # Commands
        # ==========================

        "show all projects": "all",
        "show all": "all",
        "all projects": "all",
        "all": "all",
    }

    if text in aliases:
        return aliases[text]
        # Partial Match
    for key, value in aliases.items():
        if key in text:
            return value

    return text

# =====================================================
# Page Settings
# =====================================================
st.set_page_config(
    page_title="SmartPay Project Dashboard",
    page_icon="💳",
    layout="wide"
)


st.markdown("""
<style>

/* ===========================
   Main App
=========================== */

.stApp{
    background:#F8F9FA;
}

/* ===========================
   Headings
=========================== */

h1,h2,h3,h4,h5,h6,
[data-testid="stHeading"],
[data-testid="stHeading"] *{
    color:#006747 !important;
    font-weight:700 !important;
}

/* ===========================
   Normal Text
=========================== */

p,
span{
    color:inherit !important;
}

/* ===========================
   Field Labels
=========================== */

label,
label p,
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] p{
    color:#006747 !important;
    font-weight:bold !important;
    font-size:16px !important;
}

/* ===========================
   Sidebar
=========================== */

/* ==========================================
   PREMIUM SIDEBAR
========================================== */

section[data-testid="stSidebar"]{
    background:linear-gradient(180deg,#004B34,#006747,#008A5A) !important;
    border-right:3px solid #D4AF37;
}

section[data-testid="stSidebar"] *{
    color:white !important;
    font-family:"Segoe UI",sans-serif !important;
}

/* Navigation Title */

section[data-testid="stSidebar"] label{
    color:#FFD700 !important;
    font-size:18px !important;
    font-weight:700 !important;
}

/* Radio Buttons */

div[role="radiogroup"] label{
    background:rgba(255,255,255,.08);
    margin-bottom:8px;
    padding:10px;
    border-radius:10px;
    transition:.3s;
}

div[role="radiogroup"] label:hover{
    background:rgba(255,255,255,.20);
}

/* Selected Page */

div[role="radiogroup"] label[data-selected="true"]{
    background:white !important;
    color:#006747 !important;
    font-weight:bold !important;
    border-left:5px solid #FFD700;
}

/* ===========================
   KPI Cards
=========================== */

div[data-testid="metric-container"]{
    background:white !important;
    border-left:6px solid #006747 !important;
    border-radius:12px !important;
    padding:18px !important;
    box-shadow:0 2px 8px rgba(0,0,0,.15);
}

[data-testid="stMetricLabel"]{
    color:#006747 !important;
    font-weight:bold !important;
}

[data-testid="stMetricValue"]{
    color:#222222 !important;
    font-size:34px !important;
    font-weight:700 !important;
}

[data-testid="stMetricDelta"]{
    color:#006747 !important;
}

/* ===========================
   Buttons
=========================== */

.stButton>button{
    background:#006747 !important;
    color:white !important;
    border:none !important;
    border-radius:8px !important;
    font-weight:bold !important;
}

.stButton>button:hover{
    background:#008A5A !important;
}

/* ===========================
   Text Input
=========================== */

.stTextInput input{
    background:white !important;
    color:black !important;
    border:2px solid #006747 !important;
    border-radius:8px !important;
}

/* ===========================
   Select Box
=========================== */

div[data-baseweb="select"]>div{
    background:white !important;
    color:black !important;
    border:2px solid #006747 !important;
    border-radius:8px !important;
}

/* ===========================
   DataFrame
=========================== */

[data-testid="stDataFrame"]{
    border:2px solid #006747 !important;
    border-radius:10px !important;
}

/* ===========================
   Alert Boxes
=========================== */

[data-testid="stAlert"]{
    border-radius:10px !important;
}

[data-testid="stAlert"] *{
    color:#222222 !important;
}

</style>
""", unsafe_allow_html=True)
# =====================================================
# Load Data
# =====================================================
df = load_data()

# Clean Status
df["Status"] = (
    df["Status"]
    .astype(str)
    .str.strip()
    .replace({
        "Under development": "Under Development",
        "InProcess": "In Process",
        "LIVE ": "LIVE"
    })
)

# =====================================================
# SIDEBAR
# =====================================================

st.markdown("""
<style>

/* =========================
   Sidebar
========================= */

section[data-testid="stSidebar"]{
    background:#006747;
}

/* Navigation title */
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3,
section[data-testid="stSidebar"] p{
    color:white;
}

/* Radio Labels */
div[role="radiogroup"] label{
    background:transparent !important;
    border:none !important;
    border-radius:10px;
    padding:10px 12px;
    margin-bottom:6px;
    color:white !important;
    font-weight:600;
    transition:.2s;
}

/* Hover */
div[role="radiogroup"] label:hover{
    background:#0B8758 !important;
}

/* Selected */
div[role="radiogroup"] label:has(input:checked){
    background:#0FA968 !important;
    color:white !important;
}

/* Hide Radio Circle */
div[role="radiogroup"] input{
    display:none !important;
}

/* Hide Empty Header Space */
section[data-testid="stSidebar"] div[data-testid="stVerticalBlock"]{
    padding-top:.4rem;
}

</style>
""", unsafe_allow_html=True)

# =====================================================
# LOGO
# =====================================================

st.sidebar.image(
    "smartpay_logo/smartpay_logo.png",
    use_container_width=True
)

# =====================================================
# TITLE
# =====================================================

st.sidebar.markdown("""
<h2 style="
text-align:center;
color:white;
margin-bottom:0;">
🏦 SmartPay
</h2>

<p style="
text-align:center;
color:#D1FAE5;
margin-top:2px;
margin-bottom:8px;">
Project Dashboard
</p>
""", unsafe_allow_html=True)

st.sidebar.markdown("---")

# =====================================================
# NAVIGATION
# =====================================================

st.sidebar.markdown(
    "<h4 style='color:white;'>Navigation</h4>",
    unsafe_allow_html=True
)

page = st.sidebar.radio(
    "Menu",
    [
        "Dashboard",
        "Projects",
        "Analytics",
        "Project Timeline",
        "BAU Monitoring",
        "Export"
        #"Team Performance",
        #"Voice Search",
    ],
    label_visibility="collapsed"
)

st.sidebar.markdown("---")
# ==========================================
# RESET SEARCH + SCROLL WHEN PAGE CHANGES
# ==========================================

if "previous_page" not in st.session_state:
    st.session_state.previous_page = page

elif st.session_state.previous_page != page:

    # Search / filter reset
    for key in [
        "project_search",
        "project",
        "search",
        "search_project",
        "selected_project",
        "selected_member"
    ]:
        if key in st.session_state:
            del st.session_state[key]

    # New page remember
    st.session_state.previous_page = page

    # Scroll to top
    # Scroll position reset
    st.markdown("""
    <script>
    window.parent.scrollTo(0, 0);
    window.parent.document.documentElement.scrollTop = 0;
    window.parent.document.body.scrollTop = 0;
    </script>
    """, unsafe_allow_html=True)

    st.rerun()

# =====================================================
# INFORMATION
# =====================================================

st.sidebar.markdown("""
<div style="
background:rgba(255,255,255,.12);
padding:12px;
border-radius:12px;
color:white;
font-size:14px;
line-height:1.5;">

<b>Department</b><br>
Digital Banking Group

<hr style="margin:8px 0;border:.5px solid rgba(255,255,255,.25);">

<b>Organization</b><br>
National Bank of Pakistan

<hr style="margin:8px 0;border:.5px solid rgba(255,255,255,.25);">

<b>Version</b><br>
2.0

<hr style="margin:8px 0;border:.5px solid rgba(255,255,255,.25);">

<b>Developer</b><br>
Muhammad Saad Asif

</div>
""", unsafe_allow_html=True)

st.sidebar.markdown("<br>", unsafe_allow_html=True)

st.sidebar.caption(
    "© 2026 SmartPay Dashboard"
)
# =====================================================
# DASHBOARD
# =====================================================

if page == "Dashboard":
    
    # ==========================================
    # HEADER - VIP ENTERPRISE
    # ==========================================

    city, temp, weather = get_weather()

    left, center, right = st.columns([5.5, 2.2, 2])

    # ==========================================
    # LEFT - TITLE
    # ==========================================

    with left:

        st.markdown("""
        <div style="
        background:linear-gradient(135deg,#ffffff,#f8fbff);
        border-radius:24px;
        padding:35px;
        border:1px solid #E5E7EB;
        box-shadow:0 14px 35px rgba(0,0,0,.10);
        min-height:230px;">

        <div style="
        color:#006747;
        font-size:18px;
        font-weight:700;
        letter-spacing:2px;">
        NATIONAL BANK OF PAKISTAN
        </div>

        <div style="
        font-size:52px;
        color:#006747;
        font-weight:800;
        line-height:1.15;
        margin-top:18px;">
        SmartPay Project Dashboard
        </div>

        <div style="
        margin-top:20px;
        font-size:20px;
        color:#374151;">
        Digital Banking Group
        </div>

        <div style="
        margin-top:28px;
        display:inline-block;
        background:#ECFDF5;
        color:#006747;
        padding:8px 18px;
        border-radius:30px;
        font-size:14px;
        font-weight:600;">
        ● LIVE Dashboard
        </div>

        </div>
        """, unsafe_allow_html=True)

    # ==========================================
    # CENTER - WEATHER
    # ==========================================

    with center:

        st.markdown(f"""
        <div style="
        background:linear-gradient(180deg,#ffffff,#f8fbff);
        border-radius:24px;
        padding:25px;
        min-height:230px;
        text-align:center;
        border:1px solid #E5E7EB;
        box-shadow:0 14px 35px rgba(0,0,0,.10);">

        <div style="
        font-size:32px;">
        🌤
        </div>

        <div style="
        font-size:20px;
        color:#006747;
        font-weight:700;
        margin-top:5px;">
        {city}
        </div>

        <div style="
        font-size:58px;
        color:#006747;
        font-weight:800;
        margin:18px 0 5px;">
        {temp}°
        </div>

        <div style="
        color:#6B7280;
        font-size:18px;">
        {weather}
        </div>

        </div>
        """, unsafe_allow_html=True)

    # ==========================================
    # RIGHT - DATE & TIME
    # ==========================================

    with right:

        components.html("""
        <div style="
        background:linear-gradient(180deg,#ffffff,#f8fbff);
        border-radius:24px;
        padding:25px;
        width:109%;
        min-height:230px;
        text-align:center;
        border:1px solid #E5E7EB;
        box-shadow:0 14px 35px rgba(0,0,0,.10);
        font-family:Segoe UI;
        box-sizing:border-box;">

        <div style="
        color:#006747;
        font-size:18px;
        font-weight:700;">
        TODAY
        </div>

        <div id="date"
        style="
        margin-top:18px;
        font-size:34px;
        font-weight:700;
        color:#111827;">
        </div>

        <hr style="margin:22px 0;">

        <div id="clock"
        style="
        font-size:34px;
        font-weight:800;
        color:#006747;">
        </div>

        <div style="
        margin-top:15px;
        color:#6B7280;
        font-size:14px;">
        Pakistan Standard Time
        </div>

        </div>

        <script>

        function updateClock(){

            const now = new Date();

            document.getElementById("date").innerHTML =
            now.toLocaleDateString("en-GB",{
                day:"2-digit",
                month:"short",
                year:"numeric"
            });

            document.getElementById("clock").innerHTML =
            now.toLocaleTimeString("en-US");

        }

        updateClock();

        setInterval(updateClock,1000);

        </script>
        """, height=440, scrolling=False)
    # ==========================================
    # KPI - VIP ENTERPRISE CARDS
    # ==========================================

    status = df["Status"].astype(str).str.upper().str.strip()

    total_projects = len(df)
    scoping_projects = len(df[status == "UNDER SCOPING"])
    development_projects = len(df[status == "UNDER DEVELOPMENT"])
    uat_projects = len(df[status == "UAT"])
    review_projects = len(df[status == "IS REVIEW"])
    cmc_projects = len(df[status == "CMC"])
    live_projects = len(df[status == "LIVE"])
    bau_projects = len(df[status == "BAU"])

    # Better spacing for cards
    c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(
        [1.2, 1.1, 1.2, 1.0, 1.1, 1.0, 1.0, 1.0]
    )

    ## ------------------------------------------
    # KPI Card Function
    # ------------------------------------------

    def vip_card(title, value, color):

        st.markdown(
            f"""
    <div style="
    background:white;
    border-radius:22px;
    padding:24px 16px;
    height:145px;
    border-top:8px solid {color};
    box-shadow:0 8px 22px rgba(0,0,0,.08);
    display:flex;
    flex-direction:column;
    justify-content:space-between;
    align-items:center;
    text-align:center;">

    <div style="
    color:#6B7280;
    font-size:15px;
    font-weight:600;">
    {title}
    </div>

    <div style="
    color:{color};
    font-size:48px;
    font-weight:700;
    line-height:1;">
    {value}
    </div>

    </div>
    """,
            unsafe_allow_html=True,
        )

    # ------------------------------------------
    # KPI Cards
    # ------------------------------------------

    with c1:
        vip_card("Total Projects", total_projects, "#006747")

    with c2:
        vip_card("Scoping", scoping_projects, "#8E24AA")

    with c3:
        vip_card("UAT", uat_projects, "#F9A825")

    with c4:
        vip_card("IS Review", review_projects, "#00ACC1")

    with c5:
        vip_card("CMC", cmc_projects, "#3949AB")

    with c6:
        vip_card("LIVE", live_projects, "#00C853")
    with c7:
        vip_card("BAU", bau_projects, "#607D8B")

    st.markdown("<br>", unsafe_allow_html=True)
    # =====================================================
    # TEAM OVERVIEW
    # =====================================================

    st.markdown("""
    <h2 style="
    color:#006747;
    font-size:34px;
    font-weight:700;
    margin-bottom:25px;">
    👥 Team Overview
    </h2>
    """, unsafe_allow_html=True)

    allocation = (
        df["Allocation"]
        .value_counts()
        .reset_index()
    )

    allocation.columns = ["Allocation", "Projects"]

    cols = st.columns(4)

    for i, row in allocation.iterrows():

        with cols[i % 4]:

            st.markdown(f"""
    <div style="
    background:linear-gradient(180deg,#ffffff,#f7f9fc);
    border-radius:24px;
    padding:25px;
    text-align:center;
    border:1px solid #E5E7EB;
    box-shadow:0 12px 30px rgba(0,0,0,.08);
    min-height:210px;">

    <div style="
    width:70px;
    height:70px;
    border-radius:50%;
    background:#E8F5E9;
    margin:auto;
    display:flex;
    align-items:center;
    justify-content:center;
    font-size:34px;">
    👤
    </div>

    <div style="
    margin-top:18px;
    font-size:22px;
    font-weight:700;
    color:#006747;">
    {row["Allocation"]}
    </div>

    <div style="
    margin-top:18px;
    font-size:52px;
    font-weight:800;
    color:#111827;">
    {row["Projects"]}
    </div>

    <div style="
    margin-top:8px;
    font-size:15px;
    color:#6B7280;">
    Projects Assigned
    </div>

    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)


    # =====================================================
    # TEAM WORKLOAD
    # =====================================================

    st.markdown("""
    <h2 style="
    color:#006747;
    font-size:34px;
    font-weight:700;
    margin-bottom:20px;">
    📊 Team Workload
    </h2>
    """, unsafe_allow_html=True)

    team_df = (
        df.groupby("Allocation")
        .size()
        .reset_index(name="Projects")
        .sort_values("Projects", ascending=False)
    )

    fig = px.bar(
        team_df,
        x="Allocation",
        y="Projects",
        text="Projects",
        color="Projects",
        color_continuous_scale="Greens"
    )

    fig.update_traces(
        textposition="outside",
        marker_line_width=0
    )

    fig.update_layout(
        height=500,
        plot_bgcolor="white",
        paper_bgcolor="white",
        margin=dict(l=20, r=20, t=20, b=20),
        coloraxis_showscale=False,
        xaxis_title="",
        yaxis_title="Projects",
        font=dict(size=15),
        xaxis=dict(showgrid=False),
        yaxis=dict(gridcolor="#ECECEC")
    )

    st.markdown("""
    <div style="
    background:white;
    border-radius:24px;
    padding:15px;
    border:1px solid #E5E7EB;
    box-shadow:0 12px 30px rgba(0,0,0,.08);">
    </div>
    """, unsafe_allow_html=True)

    st.plotly_chart(fig, width="stretch")


    # =====================================================
    # TEAM SUMMARY - VIP
    # =====================================================

    st.markdown("""
    <h2 style="
    color:#006747;
    font-size:34px;
    font-weight:700;
    margin-top:30px;
    margin-bottom:20px;">
    📋 Team Summary
    </h2>
    """, unsafe_allow_html=True)


    summary = (
        df.groupby("Allocation")
        .agg(
            Total=("Mandate", "count"),
            Live=("Status", lambda x: (x.str.upper() == "LIVE").sum()),
            UAT=("Status", lambda x: (x.str.upper() == "UAT").sum())
        )
        .reset_index()
    )


    # -----------------------------
    # TEAM SUMMARY TABLE CSS
    # -----------------------------

    st.markdown("""
    <style>

    .vip-summary-wrapper {
        background: #FFFFFF;
        border: 1px solid #DDE5E1;
        border-radius: 18px;
        padding: 6px;
        box-shadow: 0 8px 25px rgba(0,103,71,0.08);
        overflow: hidden;
    }

    .vip-summary-table {
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        font-size: 14px;
    }

    .vip-summary-table thead th {
        background: #006747;
        color: #FFFFFF;
        padding: 15px 16px;
        font-weight: 700;
        text-align: left;
    }

    .vip-summary-table thead th:first-child {
        border-top-left-radius: 12px;
    }

    .vip-summary-table thead th:last-child {
        border-top-right-radius: 12px;
    }

    .vip-summary-table tbody td {
        padding: 15px 16px;
        color: #1F2937;
        border-bottom: 1px solid #E5E7EB;
        background: #FFFFFF;
    }

    .vip-summary-table tbody tr:nth-child(even) td {
        background: #F8FAFC;
    }

    .vip-summary-table tbody tr:hover td {
        background: #ECFDF5;
    }

    .team-name {
        color: #006747 !important;
        font-weight: 800;
    }

    .total-count {
        background: #F0FDF4;
        color: #006747 !important;
        font-weight: 800;
        padding: 5px 10px;
        border-radius: 20px;
    }

    .live-count {
        background: #DCFCE7;
        color: #166534 !important;
        font-weight: 800;
        padding: 5px 10px;
        border-radius: 20px;
    }

    .uat-count {
        background: #FEF3C7;
        color: #92400E !important;
        font-weight: 800;
        padding: 5px 10px;
        border-radius: 20px;
    }

    </style>
    """, unsafe_allow_html=True)
  


    # =====================================================
    # CREATE SUMMARY
    # =====================================================

    summary = (
        df.groupby("Allocation")
        .agg(

            # Total Projects
            Total=("Mandate", "count"),

            # SCOPING
            Scoping=("Status", lambda x:
                x.astype(str)
                .str.strip()
                .str.upper()
                .isin([
                    "SCOPING",
                    "UNDER SCOPING"
                ])
                .sum()
            ),

            # DEVELOPMENT
            Development=("Status", lambda x:
                x.astype(str)
                .str.strip()
                .str.upper()
                .isin([
                    "DEVELOPMENT",
                    "UNDER DEVELOPMENT",
                    "SIT"
                ])
                .sum()
            ),

            # UAT
            UAT=("Status", lambda x:
                (
                    x.astype(str)
                    .str.strip()
                    .str.upper()
                    == "UAT"
                ).sum()
            ),

            # IS REVIEW
            IS_Review=("Status", lambda x:
                (
                    x.astype(str)
                    .str.strip()
                    .str.upper()
                    == "IS REVIEW"
                ).sum()
            ),

            # CMC
            CMC=("Status", lambda x:
                (
                    x.astype(str)
                    .str.strip()
                    .str.upper()
                    == "CMC"
                ).sum()
            ),

            # LIVE
            Live=("Status", lambda x:
                (
                    x.astype(str)
                    .str.strip()
                    .str.upper()
                    == "LIVE"
                ).sum()
            )
        )
        .reset_index()
    )


    # =====================================================
    # RENAME COLUMN
    # =====================================================

    summary.rename(
        columns={
            "IS_Review": "IS Review"
        },
        inplace=True
    )


    # =====================================================
    # EXACT COLUMN ORDER
    # =====================================================

    summary = summary[
        [
            "Allocation",
            "Total",
            "Scoping",
            "UAT",
            "IS Review",
            "CMC",
            "Live"
        ]
    ]


    # =====================================================
    # FORMAT SUMMARY
    # =====================================================

    summary_display = summary.copy()


    # Allocation
    summary_display["Allocation"] = summary_display[
        "Allocation"
    ].apply(
        lambda x:
        f'<span class="team-name">👤 {x}</span>'
    )


    # Total
    summary_display["Total"] = summary_display[
        "Total"
    ].apply(
        lambda x:
        f'<span class="total-count">{x}</span>'
    )


    # Scoping
    summary_display["Scoping"] = summary_display[
        "Scoping"
    ].apply(
        lambda x:
        f'<span class="scoping-count">🟠 {x}</span>'
    )





    # UAT
    summary_display["UAT"] = summary_display[
        "UAT"
    ].apply(
        lambda x:
        f'<span class="uat-count">🟡 {x}</span>'
    )


    # IS Review
    summary_display["IS Review"] = summary_display[
        "IS Review"
    ].apply(
        lambda x:
        f'<span class="review-count">🔷 {x}</span>'
    )


    # CMC
    summary_display["CMC"] = summary_display[
        "CMC"
    ].apply(
        lambda x:
        f'<span class="cmc-count">🟣 {x}</span>'
    )


    # Live
    summary_display["Live"] = summary_display[
        "Live"
    ].apply(
        lambda x:
        f'<span class="live-count">🟢 {x}</span>'
    )


    # =====================================================
    # CREATE HTML TABLE
    # =====================================================

    summary_html = summary_display.to_html(
        index=False,
        escape=False,
        classes="vip-summary-table"
    )


    # =====================================================
    # DISPLAY VIP TABLE
    # =====================================================

    st.markdown(
        f"""
        <div class="vip-summary-wrapper">
            {summary_html}
        </div>
        """,
        unsafe_allow_html=True
    )


    st.markdown("<br>", unsafe_allow_html=True)
    # =====================================================
    # SMART SEARCH - VIP PROJECT TABLE
    # =====================================================

    st.markdown("""
    <h2 style="
    color:#006747;
    font-size:34px;
    font-weight:700;
    margin-top:35px;
    margin-bottom:20px;">
    🔍 Smart Search
    </h2>
    """, unsafe_allow_html=True)


    project = st.text_input(
        "Search Project",
        placeholder="🔍 Search Project...",
        label_visibility="collapsed",
        key="project_search"
    )


    if project:

        result = df[
            df["Mandate"].astype(str).str.contains(
                project,
                case=False,
                na=False
            )
        ]


        if result.empty:

            st.error("❌ Project not found.")


        else:

            st.markdown(
                f"""
                <div style="
                background:#ECFDF5;
                border:1px solid #A7F3D0;
                border-left:6px solid #006747;
                padding:14px 18px;
                border-radius:12px;
                margin-bottom:15px;
                color:#006747;
                font-weight:700;
                font-size:15px;">
                ✅ {len(result)} Project(s) Found
                </div>
                """,
                unsafe_allow_html=True
            )


            # =================================================
            # VIP SEARCH RESULT TABLE
            # =================================================

            st.markdown("""
            <div style="
            background:#FFFFFF;
            border-radius:18px;
            padding:6px;
            border:1px solid #DDE5E1;
            box-shadow:0 8px 25px rgba(0,103,71,0.08);
            margin-bottom:20px;">
            """, unsafe_allow_html=True)


            st.dataframe(
                result,
                width="stretch",
                hide_index=True,
                height=400
            )


            st.markdown(
                "</div>",
                unsafe_allow_html=True
            )


            # =================================================
            # SEARCH SUMMARY
            # =================================================

            total_found = len(result)

            live_found = len(
                result[
                    result["Status"]
                    .astype(str)
                    .str.upper()
                    == "LIVE"
                ]
            )

            uat_found = len(
                result[
                    result["Status"]
                    .astype(str)
                    .str.upper()
                    == "UAT"
                ]
            )

            c1, c2, c3 = st.columns(3)


            with c1:
                st.metric(
                    "📋 Projects Found",
                    total_found
                )


            with c2:
                st.metric(
                    "🟢 Live",
                    live_found
                )


            with c3:
                st.metric(
                    "🟡 UAT",
                    uat_found
                )


    st.markdown("<br>", unsafe_allow_html=True)
# =====================================================
# PROJECTS
# =====================================================

elif page == "Projects":

    # =====================================================
    # HEADER
    # =====================================================

    st.markdown("""
    <div style="
    background:linear-gradient(180deg,#ffffff,#f8fbff);
    border-radius:25px;
    padding:30px;
    border:1px solid #E5E7EB;
    box-shadow:0 12px 35px rgba(0,0,0,.08);">

    <h1 style="
    color:#006747;
    margin:0;
    font-size:42px;
    font-weight:700;">
    📁 Project Portfolio
    </h1>

    <p style="
    margin-top:10px;
    color:#6B7280;
    font-size:18px;">
    SmartPay Project Management System
    </p>

    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # =====================================================
    # FILTERS
    # =====================================================

    st.markdown("""
    <h2 style="
    color:#006747;
    font-size:28px;
    font-weight:700;">
    🎯 Filters
    </h2>
    """, unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)

    with c1:
        search = st.text_input(
            "Search Project",
            placeholder="🔍 Search Project..."
        )

    with c2:
        allocation = st.selectbox(
            "Team Member",
            ["All"] + sorted(df["Allocation"].dropna().unique())
        )

    with c3:
        status = st.selectbox(
            "Status",
            ["All"] + sorted(df["Status"].dropna().unique())
        )

    with c4:
        category = st.selectbox(
            "Category",
            ["All"] + sorted(df["Category"].dropna().unique())
        )

    # =====================================================
    # FILTERING
    # =====================================================

    filtered_df = df.copy()

    if search:
        filtered_df = filtered_df[
            filtered_df["Mandate"].astype(str).str.contains(
                search,
                case=False,
                na=False
            )
        ]

    if allocation != "All":
        filtered_df = filtered_df[
            filtered_df["Allocation"] == allocation
        ]

    if status != "All":
        filtered_df = filtered_df[
            filtered_df["Status"] == status
        ]

    if category != "All":
        filtered_df = filtered_df[
            filtered_df["Category"] == category
        ]

    st.markdown("<br>", unsafe_allow_html=True)

    # =====================================================
    # SUMMARY
    # =====================================================

    left, right = st.columns([3,1])

    with left:

        st.markdown(f"""
        <div style="
        background:#E8F5E9;
        padding:18px;
        border-radius:15px;
        color:#006747;
        font-size:20px;
        font-weight:700;">
        📌 Showing <b>{len(filtered_df)}</b> Project(s)
        </div>
        """, unsafe_allow_html=True)

    with right:

        csv = filtered_df.to_csv(index=False).encode("utf-8")

        st.download_button(
            "⬇ Download CSV",
            csv,
            "Projects.csv",
            "text/csv",
            width="stretch"
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # =====================================================
    # VIP CLICKABLE KPI CARDS
    # =====================================================

    # -----------------------------------------
    # STATUS CLEAN
    # -----------------------------------------

    status_clean = (
        filtered_df["Status"]
        .astype(str)
        .str.strip()
        .str.upper()
    )


    # -----------------------------------------
    # COUNTS
    # -----------------------------------------

    all_count = len(filtered_df)

    scoping_count = status_clean.isin(
        ["SCOPING", "UNDER SCOPING"]
    ).sum()

    development_count = status_clean.isin(
        ["DEVELOPMENT", "UNDER DEVELOPMENT", "SIT"]
    ).sum()

    uat_count = (
        status_clean == "UAT"
    ).sum()

    review_count = (
        status_clean == "IS REVIEW"
    ).sum()

    cmc_count = (
        status_clean == "CMC"
    ).sum()

    live_count = (
        status_clean == "LIVE"
    ).sum()

    bau_count = (
        status_clean == "BAU"
    ).sum()


    # =====================================================
    # VIP KPI CARD CSS
    # =====================================================

    st.markdown("""
    <style>

    /* =====================================================
    MAIN KPI CONTAINER
    ===================================================== */

    .st-key-status_kpis {
        width:100%;
    }


    /* =====================================================
    KPI COLUMNS
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"] {

        gap:12px !important;

    }


    /* =====================================================
    KPI CARD
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stButton"] {

        width:100% !important;

    }


    .st-key-status_kpis
    div[data-testid="stButton"] button {

        width:100% !important;

        min-height:145px !important;

        background:#FFFFFF !important;

        border:1px solid #E5E7EB !important;

        border-radius:20px !important;

        padding:18px 10px !important;

        box-shadow:
            0 8px 24px rgba(0,0,0,0.07) !important;

        color:#374151 !important;

        font-family:
            "Segoe UI",
            Arial,
            sans-serif !important;

        font-size:15px !important;

        font-weight:600 !important;

        line-height:1.7 !important;

        white-space:pre-line !important;

        text-align:center !important;

        transition:
            transform 0.18s ease,
            box-shadow 0.18s ease,
            background 0.18s ease !important;

    }


    /* =====================================================
    HOVER
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stButton"] button:hover {

        background:#FFFFFF !important;

        transform:translateY(-3px);

        box-shadow:
            0 14px 30px rgba(0,0,0,0.10) !important;

    }


    /* =====================================================
    FOCUS
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stButton"] button:focus {

        outline:none !important;

    }


    /* =====================================================
    BUTTON TEXT
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stButton"] button p {

        margin:0 !important;

        padding:0 !important;

        font-family:
            "Segoe UI",
            Arial,
            sans-serif !important;

        font-size:15px !important;

        font-weight:600 !important;

        color:#374151 !important;

        line-height:1.9 !important;

    }


    /* =====================================================
    ALL
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"]:nth-child(1)
    div[data-testid="stColumn"]:nth-child(1)
    div[data-testid="stButton"] button {

        border-top:7px solid #006747 !important;

    }


    /* =====================================================
    SCOPING
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"]:nth-child(1)
    div[data-testid="stColumn"]:nth-child(2)
    div[data-testid="stButton"] button {

        border-top:7px solid #8E24AA !important;

    }


    /* =====================================================
    DEVELOPMENT
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"]:nth-child(1)
    div[data-testid="stColumn"]:nth-child(3)
    div[data-testid="stButton"] button {

        border-top:7px solid #FF9800 !important;

    }


    /* =====================================================
    UAT
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"]:nth-child(1)
    div[data-testid="stColumn"]:nth-child(4)
    div[data-testid="stButton"] button {

        border-top:7px solid #F9A825 !important;

    }


    /* =====================================================
    IS REVIEW
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"]:nth-child(1)
    div[data-testid="stColumn"]:nth-child(5)
    div[data-testid="stButton"] button {

        border-top:7px solid #00ACC1 !important;

    }


    /* =====================================================
    CMC
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"]:nth-child(1)
    div[data-testid="stColumn"]:nth-child(6)
    div[data-testid="stButton"] button {

        border-top:7px solid #3949AB !important;

    }


    /* =====================================================
    LIVE
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"]:nth-child(1)
    div[data-testid="stColumn"]:nth-child(7)
    div[data-testid="stButton"] button {

        border-top:7px solid #00C853 !important;

    }


    /* =====================================================
    BAU
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stHorizontalBlock"]:nth-child(1)
    div[data-testid="stColumn"]:nth-child(8)
    div[data-testid="stButton"] button {

        border-top:7px solid #607D8B !important;

    }


    /* =====================================================
    REMOVE EXTRA STREAMLIT SPACING
    ===================================================== */

    .st-key-status_kpis
    div[data-testid="stVerticalBlock"] {

        gap:0 !important;

    }

    </style>
    """, unsafe_allow_html=True)

    # =====================================================
    # KPI CARD CONTAINER
    # =====================================================

    with st.container(key="status_kpis"):

        c1, c2, c3, c4, c5, c6, c7 = st.columns(7)


        # -----------------------------------------
        # ALL
        # -----------------------------------------

        with c1:

            if st.button(
                f"ALL\n\n{all_count}",
                key="kpi_all",
                use_container_width=True
            ):
                st.session_state["project_status_filter"] = "ALL"
                st.rerun()


        # -----------------------------------------
        # SCOPING
        # -----------------------------------------

        with c2:

            if st.button(
                f"SCOPING\n\n{scoping_count}",
                key="kpi_scoping",
                use_container_width=True
            ):
                st.session_state["project_status_filter"] = "SCOPING"
                st.rerun()


        

        # -----------------------------------------
        # UAT
        # -----------------------------------------

        with c3:

            if st.button(
                f"UAT\n\n{uat_count}",
                key="kpi_uat",
                use_container_width=True
            ):
                st.session_state["project_status_filter"] = "UAT"
                st.rerun()


        # -----------------------------------------
        # IS REVIEW
        # -----------------------------------------

        with c4:

            if st.button(
                f"IS REVIEW\n\n{review_count}",
                key="kpi_review",
                use_container_width=True
            ):
                st.session_state["project_status_filter"] = "IS REVIEW"
                st.rerun()


        # -----------------------------------------
        # CMC
        # -----------------------------------------

        with c5:

            if st.button(
                f"CMC\n\n{cmc_count}",
                key="kpi_cmc",
                use_container_width=True
            ):
                st.session_state["project_status_filter"] = "CMC"
                st.rerun()


        # -----------------------------------------
        # LIVE
        # -----------------------------------------

        with c6:

            if st.button(
                f"LIVE\n\n{live_count}",
                key="kpi_live",
                use_container_width=True
            ):
                st.session_state["project_status_filter"] = "LIVE"
                st.rerun()


        # -----------------------------------------
        # BAU
        # -----------------------------------------

        with c7:

            if st.button(
                f"BAU\n\n{bau_count}",
                key="kpi_bau",
                use_container_width=True
            ):
                st.session_state["project_status_filter"] = "BAU"
                st.rerun()


    st.markdown("<br>", unsafe_allow_html=True)
    # =====================================================
    # APPLY KPI STATUS FILTER TO EXISTING TABLE
    # =====================================================

    selected_status = st.session_state.get(
        "project_status_filter",
        "ALL"
    )

    status_upper = (
        filtered_df["Status"]
        .astype(str)
        .str.strip()
        .str.upper()
    )


    if selected_status == "SCOPING":

        filtered_df = filtered_df[
            status_upper.isin([
                "SCOPING",
                "UNDER SCOPING"
            ])
        ].copy()


    elif selected_status == "DEVELOPMENT":

        filtered_df = filtered_df[
            status_upper.isin([
                "DEVELOPMENT",
                "UNDER DEVELOPMENT",
                "SIT"
            ])
        ].copy()


    elif selected_status == "UAT":

        filtered_df = filtered_df[
            status_upper == "UAT"
        ].copy()


    elif selected_status == "IS REVIEW":

        filtered_df = filtered_df[
            status_upper == "IS REVIEW"
        ].copy()


    elif selected_status == "CMC":

        filtered_df = filtered_df[
            status_upper == "CMC"
        ].copy()


    elif selected_status == "LIVE":

        filtered_df = filtered_df[
            status_upper == "LIVE"
        ].copy()


    elif selected_status == "BAU":

        filtered_df = filtered_df[
            status_upper == "BAU"
        ].copy()
    # =====================================================
    # PROJECT TABLE - VIP STYLE
    # =====================================================

    st.markdown("""
    <h2 style="
    color:#006747;
    font-size:30px;
    font-weight:700;
    margin-bottom:18px;">
    📋 Project Details
    </h2>
    """, unsafe_allow_html=True)


    # -----------------------------
    # VIP TABLE CSS
    # -----------------------------

    st.markdown("""
    <style>

    .project-table-wrapper {
        background: #FFFFFF;
        border: 1px solid #DDE5E1;
        border-radius: 16px;
        padding: 6px;
        box-shadow: 0 6px 20px rgba(0,103,71,0.08);
        overflow: hidden;
    }

    .project-table {
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        font-size: 14px;
    }

    .project-table thead th {
        background: #006747;
        color: white;
        font-weight: 700;
        padding: 14px 12px;
        text-align: left;
        border: none;
    }

    .project-table thead th:first-child {
        border-top-left-radius: 11px;
    }

    .project-table thead th:last-child {
        border-top-right-radius: 11px;
    }

    .project-table tbody td {
        padding: 13px 12px;
        color: #1F2937;
        border-bottom: 1px solid #E5E7EB;
        background: #FFFFFF;
    }

    .project-table tbody tr:nth-child(even) td {
        background: #F8FAFC;
    }

    .project-table tbody tr:hover td {
        background: #ECFDF5;
    }

    .project-name {
        font-weight: 700;
        color: #006747 !important;
    }

    .status-badge {
        display: inline-block;
        padding: 5px 11px;
        border-radius: 20px;
        font-size: 11px;
        font-weight: 800;
        white-space: nowrap;
    }

    .status-live {
        background: #DCFCE7;
        color: #166534;
    }

    .status-uat {
        background: #FEF3C7;
        color: #92400E;
    }

    .status-development {
        background: #DBEAFE;
        color: #1E40AF;
    }

    .status-review {
        background: #CCFBF1;
        color: #115E59;
    }

    .status-cmc {
        background: #EDE9FE;
        color: #5B21B6;
    }

    .status-scoping {
        background: #F3E8FF;
        color: #7E22CE;
    }

    .status-default {
        background: #F3F4F6;
        color: #374151;
    }

    </style>
    """, unsafe_allow_html=True)


    # -----------------------------
    # CREATE VIP TABLE
    # -----------------------------

    display_df = filtered_df.copy()


    # ==========================================
    # DATE FORMAT
    # ==========================================

    if "Date" in display_df.columns:

        display_df["Date"] = pd.to_datetime(
            display_df["Date"],
            errors="coerce"
        ).dt.strftime("%Y-%m-%d")


    # ==========================================
    # LIVE DATE FORMAT
    # ==========================================

    if "Live Date" in display_df.columns:

        def format_live_date(value):

            if pd.isna(value):
                return "TBD"

            value = str(value).strip()

            if value.upper() == "BAU":
                return "BAU"

            parsed = pd.to_datetime(
                value,
                errors="coerce"
            )

            if pd.notna(parsed):
                return parsed.strftime("%Y-%m-%d")

            return value


        display_df["Live Date"] = (
            display_df["Live Date"]
            .apply(format_live_date)
        )

    def status_badge(status):

        status = str(status).strip()
        status_upper = status.upper()

        if status_upper == "LIVE":
            css = "status-live"

        elif status_upper == "UAT":
            css = "status-uat"

        elif status_upper in [
            "DEVELOPMENT",
            "UNDER DEVELOPMENT",
            "SIT"
        ]:
            css = "status-development"

        elif status_upper == "IS REVIEW":
            css = "status-review"

        elif status_upper == "CMC":
            css = "status-cmc"

        elif status_upper in [
            "SCOPING",
            "UNDER SCOPING"
        ]:
            css = "status-scoping"

        else:
            css = "status-default"

        return f'<span class="status-badge {css}">{status}</span>'


    # Status ko badge mein convert karo
    if "Status" in display_df.columns:
        display_df["Status"] = display_df["Status"].apply(status_badge)


    # Project/Mandate name ko highlight karo
    if "Mandate" in display_df.columns:
        display_df["Mandate"] = display_df["Mandate"].apply(
            lambda x: f'<span class="project-name">📁 {x}</span>'
        )


    # -----------------------------
    # HTML TABLE
    # -----------------------------

    table_html = display_df.to_html(
        index=False,
        escape=False,
        classes="project-table"
    )


    st.markdown(
        f"""
        <div class="project-table-wrapper">
            {table_html}
        </div>
        """,
        unsafe_allow_html=True
    )
# =====================================================
# ANALYTICS
# =====================================================

elif page == "Analytics":
    # ==========================================
    # ANALYTICS DATA
    # ==========================================

    analytics_source_df = df[
        df["Status"]
        .astype(str)
        .str.strip()
        .str.upper()
        != "BAU"
    ].copy()

    bau_df = df[
        df["Status"]
        .astype(str)
        .str.strip()
        .str.upper()
        == "BAU"
    ].copy()
    
    st.markdown("""
    <div style="
    background:linear-gradient(180deg,#ffffff,#f8fbff);
    border-radius:24px;
    padding:28px;
    border:1px solid #E5E7EB;
    box-shadow:0 14px 35px rgba(0,0,0,.10);">

    <h1 style="
    color:#006747;
    margin:0;
    font-size:40px;
    font-weight:700;">
    📊 Analytics Dashboard
    </h1>

    <p style="
    color:#6B7280;
    font-size:17px;
    margin-top:10px;">
    SmartPay Project Insights & Team Performance
    </p>

    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    # ==========================================
    # KPI CARDS
    # ==========================================

    status = df["Status"].astype(str).str.upper().str.strip()

    total_projects = len(df)
    live_projects = len(df[status == "LIVE"])
    uat_projects = len(df[status == "UAT"])
    team_members = df["Allocation"].nunique()

    live_percent = round((live_projects / total_projects) * 100, 1) if total_projects else 0
    uat_percent = round((uat_projects / total_projects) * 100, 1) if total_projects else 0


    def analytics_card(title, value, color):

        st.markdown(f"""
        <div style="
        background:white;
        border-radius:22px;
        padding:24px;
        text-align:center;
        border-top:7px solid {color};
        box-shadow:0 10px 30px rgba(0,0,0,.10);
        min-height:140px;">

        <div style="
        color:#6B7280;
        font-size:15px;
        font-weight:600;">
        {title}
        </div>

        <div style="
        color:{color};
        font-size:46px;
        font-weight:700;
        margin-top:18px;">
        {value}
        </div>

        </div>
        """, unsafe_allow_html=True)


    k1, k2, k3, k4 = st.columns(4)

    with k1:
        analytics_card("Total Projects", total_projects, "#006747")

    with k2:
        analytics_card("Live %", f"{live_percent}%", "#00C853")

    with k3:
        analytics_card("UAT %", f"{uat_percent}%", "#F9A825")

    with k4:
        analytics_card("Team Members", team_members, "#3949AB")

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # FILTERS
    # ==========================================

    f1, f2 = st.columns([2,2])

    with f1:
        selected_allocation = st.selectbox(
            "👤 Team Member",
            ["All"] + sorted(df["Allocation"].dropna().unique())
        )

    with f2:
        selected_status = st.selectbox(
            "📌 Status",
            ["All"] + sorted(df["Status"].dropna().unique())
        )

    analytics_df = df.copy()

    if selected_allocation != "All":
        analytics_df = analytics_df[
            analytics_df["Allocation"] == selected_allocation
        ]

    if selected_status != "All":
        analytics_df = analytics_df[
            analytics_df["Status"] == selected_status
        ]

    analytics_df["Status"] = (
        analytics_df["Status"]
        .astype(str)
        .str.upper()
        .str.strip()
    )

    status_order = [
        "LIVE",
        "UAT",
        "IS REVIEW",
        "CMC",
        "UNDER SCOPING"
    ]

    color_map = {
        "LIVE":"#00C853",
        "UAT":"#F9A825",
        "IS REVIEW":"#00ACC1",
        "CMC":"#3949AB",
        "UNDER SCOPING":"#8E24AA"
    }

    # ==========================================
    # STATUS CHART
    # ==========================================

    st.markdown("""
    <h2 style="color:#006747;font-size:30px;">
    📈 Project Status
    </h2>
    """, unsafe_allow_html=True)

    status_count = (
        analytics_df["Status"]
        .value_counts()
        .reindex(status_order, fill_value=0)
        .reset_index()
    )

    status_count.columns = [
        "Status",
        "Projects"
    ]

    fig1 = px.bar(
        status_count,
        x="Status",
        y="Projects",
        text="Projects",
        color="Status",
        color_discrete_map=color_map
    )

    fig1.update_layout(
        height=470,
        template="plotly_white",
        plot_bgcolor="white",
        paper_bgcolor="white",
        title_text="",
        coloraxis_showscale=False,

        font=dict(
            color="#111827",
            size=14
        ),

        xaxis=dict(
            title="Status",
            title_font=dict(color="#374151", size=14),
            tickfont=dict(color="#374151", size=13),
            showgrid=False,
            zeroline=False
        ),

        yaxis=dict(
            title="Projects",
            title_font=dict(color="#374151", size=14),
            tickfont=dict(color="#374151", size=13),
            gridcolor="#E5E7EB",
            zeroline=False
        ),

        legend=dict(
            font=dict(
                color="#374151",
                size=12
            )
        ),

        margin=dict(
            l=50,
            r=30,
            t=20,
            b=60
        )
    )

    fig1.update_traces(
        textposition="inside",
        marker_line_width=0
    )

    st.plotly_chart(fig1, width="stretch")

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # TEAM PERFORMANCE
    # ==========================================

    st.markdown("""
    <h2 style="color:#006747;font-size:30px;">
    👥 Team Performance
    </h2>
    """, unsafe_allow_html=True)

    allocation_count = (
        analytics_df["Allocation"]
        .value_counts()
        .reset_index()
    )

    allocation_count.columns = [
        "Allocation",
        "Projects"
    ]

    fig2 = px.bar(
        allocation_count,
        x="Allocation",
        y="Projects",
        text="Projects",
        color="Projects",
        color_continuous_scale="Greens"
    )

    fig2.update_layout(
        height=470,
        template="plotly_white",
        plot_bgcolor="white",
        paper_bgcolor="white",
        coloraxis_showscale=False,

        font=dict(
            color="#111827",
            size=14
        ),

        xaxis=dict(
            title="",
            tickfont=dict(
                color="#374151",
                size=13
            ),
            showgrid=False,
            zeroline=False
        ),

        yaxis=dict(
            title="Projects",
            title_font=dict(
                color="#374151",
                size=14
            ),
            tickfont=dict(
                color="#374151",
                size=13
            ),
            gridcolor="#E5E7EB",
            zeroline=False
        ),

        margin=dict(
            l=50,
            r=30,
            t=20,
            b=80
        )
    )

    fig2.update_traces(
        textposition="outside",
        textfont=dict(
            color="#111827",
            size=13
        )
    )

    st.plotly_chart(fig2, width="stretch")

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # PERSON ANALYTICS
    # ==========================================

    if selected_allocation != "All":

        # ==========================================
        # PERSON DATA
        # ==========================================

        person_df = analytics_df[
            analytics_df["Allocation"] == selected_allocation
        ].copy()

        # ==========================================
        # PIE CHART - TOP
        # ==========================================

        st.markdown(
            f"""
            <h2 style="
            color:#006747;
            font-size:28px;
            margin-bottom:5px;">
            🥧 {selected_allocation}
            </h2>
            """,
            unsafe_allow_html=True
        )

        person_status = (
            person_df["Status"]
            .value_counts()
            .reindex(status_order, fill_value=0)
            .reset_index()
        )

        person_status.columns = [
            "Status",
            "Projects"
        ]

        fig3 = px.pie(
            person_status,
            names="Status",
            values="Projects",
            hole=.55,
            color="Status",
            color_discrete_map=color_map
        )

        fig3.update_layout(
            height=430,
            template="plotly_white",
            plot_bgcolor="white",
            paper_bgcolor="white",

            font=dict(
                color="#111827",
                size=13
            ),

            legend=dict(
                font=dict(
                    color="#374151",
                    size=12
                )
            ),

            margin=dict(
                l=20,
                r=20,
                t=10,
                b=10
            )
        )

        fig3.update_traces(
            textfont=dict(
                color="white",
                size=13
            )
        )

        st.plotly_chart(
            fig3,
            width="stretch"
        )

        # ==========================================
        # PROJECT STATUS PROGRESS - BELOW PIE
        # ==========================================

        st.markdown(
            """
            <h2 style="
            color:#006747;
            font-size:28px;
            margin-top:10px;
            margin-bottom:20px;">
            📊 Project Status Progress
            </h2>
            """,
            unsafe_allow_html=True
        )

        status_stages = [
            "SCOPING",
            "UAT",
            "IS REVIEW",
            "CMC",
            "LIVE"
        ]

        status_progress = {
            "SCOPING": 1,
            "UNDER SCOPING": 1,
            "UAT": 2,
            "IS REVIEW": 3,
            "CMC": 4,
            "LIVE": 5
        }

        # ==========================================
        # REMOVE BAU PROJECTS
        # ==========================================

        project_df = person_df[
            person_df["Status"].astype(str).str.strip().str.upper() != "BAU"
        ].copy()

        # ==========================================
        # SHOW ONLY NON-BAU PROJECTS
        # ==========================================

        for _, row in project_df.iterrows():

            project_name = str(row["Mandate"])
            current_status = str(row["Status"]).strip().upper()

            current_stage = status_progress.get(
                current_status,
                1
            )

            progress_percent = int(
                (current_stage / 5) * 100
            )

            # ==================================
            # STATUS COLORS
            # ==================================

            if current_status == "LIVE":

                status_bg = "#DCFCE7"
                status_color = "#166534"

            elif current_status == "UAT":

                status_bg = "#FEF3C7"
                status_color = "#92400E"


            else:

                status_bg = "#F3F4F6"
                status_color = "#374151"

            # ==================================
            # PROJECT CARD
            # ==================================

            card_html = f"""
            <div style="
            background:#FFFFFF;
            border:1px solid #E5E7EB;
            border-radius:16px;
            padding:18px 20px;
            margin-bottom:6px;
            box-shadow:0 4px 12px rgba(0,0,0,0.06);
            ">

            <div style="
            display:flex;
            justify-content:space-between;
            align-items:center;
            gap:20px;
            margin-bottom:15px;
            ">

            <div style="
            color:#111827;
            font-size:16px;
            font-weight:700;
            flex:1;
            line-height:1.4;
            word-break:break-word;
            ">
            📁 {project_name}
            </div>

            <div style="
            background:{status_bg};
            color:{status_color};
            padding:6px 12px;
            border-radius:20px;
            font-size:10px;
            font-weight:800;
            white-space:nowrap;
            flex-shrink:0;
            ">
            {current_status}
            </div>

            </div>

            <div style="
            width:100%;
            height:8px;
            background:#E5E7EB;
            border-radius:10px;
            overflow:hidden;
            ">

            <div style="
            width:{progress_percent}%;
            height:100%;
            background:#006747;
            border-radius:10px;
            ">
            </div>

            </div>

            </div>
            """

            st.html(card_html)

            # ==================================
            # STATUS STEPS
            # ==================================

            steps_html = """
            <div style="
            display:flex;
            justify-content:space-between;
            align-items:flex-start;
            margin:0 0 28px 0;
            padding:0 5px;
            ">
            """

            for i, stage in enumerate(status_stages, start=1):

                if i <= current_stage:

                    dot_color = "#006747"
                    text_color = "#006747"

                else:

                    dot_color = "#D1D5DB"
                    text_color = "#9CA3AF"

                steps_html += f"""
            <div style="
            flex:1;
            text-align:center;
            ">

            <div style="
            width:10px;
            height:10px;
            background:{dot_color};
            border-radius:50%;
            margin:auto;
            ">
            </div>

            <div style="
            margin-top:6px;
            color:{text_color};
            font-size:9px;
            font-weight:700;
            white-space:nowrap;
            ">
            {stage}
            </div>

            </div>
            """

            steps_html += """
            </div>
            """

            st.html(steps_html)
        # ==========================================
        # PROJECT DETAILS
        # ==========================================

        display_person_df = person_df.copy()


        # ------------------------------------------
        # DATE FORMAT
        # ------------------------------------------

        for col in ["Date", "Live Date"]:

            if col in display_person_df.columns:

                original = display_person_df[col].copy()

                parsed = pd.to_datetime(
                    original,
                    errors="coerce"
                )

                formatted = parsed.dt.strftime(
                    "%Y-%m-%d"
                )

                # Keep text like BAU
                display_person_df[col] = formatted.where(
                    parsed.notna(),
                    original.astype(str)
                )

                # Blank values
                display_person_df[col] = display_person_df[col].replace(
                    ["nan", "NaT", "", "None"],
                    "TBD"
                )


        st.dataframe(
            display_person_df,
            width="stretch",
            hide_index=True
        )
    # =====================================================
    # BAU MONITORING ANALYTICS
    # =====================================================

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("""
    <h2 style="
    color:#006747;
    font-size:30px;
    font-weight:700;
    margin-top:25px;">
    🏦 BAU Monitoring
    </h2>

    <p style="
    color:#6B7280;
    font-size:16px;
    margin-bottom:20px;">
    Business as Usual projects are monitored separately from delivery projects.
    </p>
    """, unsafe_allow_html=True)


    # =====================================================
    # BAU DATA
    # =====================================================

    bau_df = df[
        df["Status"]
        .astype(str)
        .str.strip()
        .str.upper()
        == "BAU"
    ].copy()


    # =====================================================
    # BAU KPI
    # =====================================================

    bau_total = len(bau_df)

    bau_owners = (
        bau_df["Allocation"]
        .dropna()
        .astype(str)
        .nunique()
    )

    bau_updates = (
        bau_df["Update"]
        .notna()
        .sum()
        if "Update" in bau_df.columns
        else 0
    )


    b1, b2, b3 = st.columns(3)


    with b1:
        analytics_card(
            "BAU Projects",
            bau_total,
            "#607D8B"
        )

    with b2:
        analytics_card(
            "BAU Owners",
            bau_owners,
            "#3949AB"
        )

    with b3:
        analytics_card(
            "BAU Updates",
            bau_updates,
            "#00897B"
        )


    st.markdown("<br>", unsafe_allow_html=True)


    # =====================================================
    # BAU PROJECTS BY OWNER
    # =====================================================

    if not bau_df.empty:

        st.markdown("""
        <h2 style="
        color:#006747;
        font-size:26px;
        font-weight:700;">
        👥 BAU Projects by Owner
        </h2>
        """, unsafe_allow_html=True)


        bau_owner_count = (
            bau_df["Allocation"]
            .astype(str)
            .value_counts()
            .reset_index()
        )

        bau_owner_count.columns = [
            "Allocation",
            "Projects"
        ]


        fig_bau = px.bar(
            bau_owner_count,
            x="Allocation",
            y="Projects",
            text="Projects",
            color="Projects",
            color_continuous_scale="Greens"
        )


        fig_bau.update_traces(
            textposition="outside",
            marker_line_width=0
        )


        fig_bau.update_layout(
            height=450,
            template="plotly_white",
            plot_bgcolor="white",
            paper_bgcolor="white",
            coloraxis_showscale=False,

            xaxis=dict(
                title="",
                showgrid=False
            ),

            yaxis=dict(
                title="BAU Projects",
                gridcolor="#E5E7EB"
            ),

            margin=dict(
                l=50,
                r=30,
                t=20,
                b=60
            )
        )


        st.plotly_chart(
            fig_bau,
            width="stretch"
        )
# =====================================================
# PROJECT TIMELINE
# =====================================================

elif page == "Project Timeline":
    # ==========================================
    # TIMELINE DATA
    # EXCLUDE BAU PROJECTS
    # ==========================================

    timeline_df = df[
        df["Status"]
        .astype(str)
        .str.strip()
        .str.upper()
        != "BAU"
    ].copy()
    # ==========================================
    # HEADER
    # ==========================================

    st.html("""
    <div style="
        background:linear-gradient(180deg,#ffffff,#f8fbff);
        border-radius:24px;
        padding:28px;
        border:1px solid #E5E7EB;
        box-shadow:0 14px 35px rgba(0,0,0,.10);
        margin-bottom:20px;
        font-family:Segoe UI,Arial,sans-serif;">

        <div style="
            color:#006747;
            font-size:14px;
            font-weight:700;
            letter-spacing:2px;
            margin-bottom:8px;">
            SMARTPAY PROJECT MANAGEMENT
        </div>

        <div style="
            color:#006747;
            font-size:40px;
            font-weight:700;
            margin:0;">
            Project Timeline
        </div>

        <div style="
            color:#6B7280;
            font-size:17px;
            margin-top:10px;">
            Track the current progress of SmartPay projects across every delivery stage.
        </div>

    </div>
    """)


    # ==========================================
    # LEGEND
    # ==========================================

    st.html("""
    <div style="
        background:white;
        border-radius:18px;
        padding:18px 22px;
        border:1px solid #E5E7EB;
        box-shadow:0 8px 20px rgba(0,0,0,.06);
        margin-bottom:25px;
        font-family:Segoe UI,Arial,sans-serif;
    ">

        <div style="
            color:#006747;
            font-size:16px;
            font-weight:700;
            margin-bottom:14px;">
            Timeline Status
        </div>

        <div style="
            display:flex;
            align-items:center;
            gap:28px;
            font-size:14px;
            font-weight:600;">

            <div>
                <span style="
                    display:inline-block;
                    width:13px;
                    height:13px;
                    background:#16A34A;
                    border-radius:50%;
                    margin-right:7px;">
                </span>
                <span style="color:#374151;">
                    Completed
                </span>
            </div>

            <div>
                <span style="
                    display:inline-block;
                    width:13px;
                    height:13px;
                    background:#F59E0B;
                    border-radius:50%;
                    margin-right:7px;">
                </span>
                <span style="color:#374151;">
                    Current Stage
                </span>
            </div>

            <div>
                <span style="
                    display:inline-block;
                    width:13px;
                    height:13px;
                    background:#D1D5DB;
                    border-radius:50%;
                    margin-right:7px;">
                </span>
                <span style="color:#374151;">
                    Pending
                </span>
            </div>

        </div>

    </div>
    """)


    # ==========================================
    # SEARCH
    # ==========================================

    st.markdown("""
    <h2 style="
        color:#006747;
        font-size:28px;
        margin-bottom:15px;">
        🔎 Find Project Timeline
    </h2>
    """, unsafe_allow_html=True)


    search_type = st.radio(
        "Search By",
        ["Project", "Team Member"],
        horizontal=True
    )


    if search_type == "Project":

        selected = st.selectbox(
        "Select Project",
        ["All Projects"] +
        sorted(
            timeline_df["Mandate"]
            .dropna()
            .astype(str)
            .unique()
        )
    )

    else:

        selected = st.selectbox(
        "Select Team Member",
        ["All Members"] +
        sorted(
            timeline_df["Allocation"]
            .dropna()
            .astype(str)
            .unique()
        )
    )


    st.markdown("<br>", unsafe_allow_html=True)


    # ==========================================
    # STAGES
    # ==========================================

    stages = [
        "SCOPING",
        "DEVELOPMENT",
        "UAT",
        "IS REVIEW",
        "CMC",
        "LIVE"
    ]


    # ==========================================
    # TIMELINE FUNCTION
    # ==========================================

    def show_timeline(project):

        current = str(project["Status"]).upper().strip()

        # STATUS MAPPING

        if current in [
            "SIT",
            "UNDER DEVELOPMENT",
            "DEVELOPMENT"
        ]:
            current = "DEVELOPMENT"

        elif current in [
            "UNDER SCOPING",
            "SCOPING"
        ]:
            current = "SCOPING"

        elif current not in stages:
            current = "SCOPING"


        current_index = stages.index(current)


        # ======================================
        # PROJECT TITLE
        # ======================================

        st.html(f"""
        <div style="
            background:linear-gradient(180deg,#ffffff,#f9fbfd);
            border-radius:22px;
            padding:25px;
            margin-bottom:12px;
            border:1px solid #E5E7EB;
            box-shadow:0 10px 28px rgba(0,0,0,.08);
            font-family:Segoe UI,Arial,sans-serif;">

            <div style="
                color:#6B7280;
                font-size:12px;
                font-weight:600;
                letter-spacing:1px;">
                PROJECT TIMELINE
            </div>

            <div style="
                color:#006747;
                font-size:25px;
                font-weight:700;
                margin-top:5px;">
                📌 {project["Mandate"]}
            </div>

        </div>
        """)


        # ======================================
        # TIMELINE HTML
        # ======================================

        timeline_html = """
        <div style="
            background:white;
            border-radius:20px;
            padding:28px 20px;
            border:1px solid #E5E7EB;
            box-shadow:0 8px 22px rgba(0,0,0,.06);
            font-family:Segoe UI,Arial,sans-serif;
            overflow-x:auto;">

            <div style="
                display:flex;
                align-items:flex-start;
                min-width:700px;">
        """


        for i, stage in enumerate(stages):

            if i < current_index:
                color = "#16A34A"
                symbol = "✓"

            elif i == current_index:
                color = "#F59E0B"
                symbol = "●"

            else:
                color = "#D1D5DB"
                symbol = "○"


            # connector
            connector = ""

            if i < len(stages) - 1:

                if i < current_index:
                    line_color = "#16A34A"
                else:
                    line_color = "#D1D5DB"

                connector = f"""
                <div style="
                    flex:1;
                    height:4px;
                    background:{line_color};
                    margin-top:18px;">
                </div>
                """


            timeline_html += f"""

            <div style="
                width:90px;
                text-align:center;
                flex-shrink:0;">

                <div style="
                    width:38px;
                    height:38px;
                    border-radius:50%;
                    background:{color};
                    color:white;
                    margin:auto;
                    line-height:38px;
                    font-size:18px;
                    font-weight:700;
                    box-shadow:0 4px 12px rgba(0,0,0,.15);">
                    {symbol}
                </div>

                <div style="
                    margin-top:10px;
                    color:#374151;
                    font-size:11px;
                    font-weight:700;">
                    {stage}
                </div>

            </div>

            {connector}
            """


        timeline_html += """
            </div>
        </div>
        """


        st.html(timeline_html)


        st.markdown("<br>", unsafe_allow_html=True)


        # ======================================
        # PROJECT INFORMATION
        # ======================================

        c1, c2, c3 = st.columns(3)


        with c1:

            st.html(f"""
            <div style="
                background:white;
                border-radius:16px;
                padding:18px;
                border:1px solid #E5E7EB;
                box-shadow:0 5px 15px rgba(0,0,0,.05);
                font-family:Segoe UI,Arial,sans-serif;">

                <div style="
                    color:#6B7280;
                    font-size:12px;
                    font-weight:600;">
                    PROJECT
                </div>

                <div style="
                    color:#111827;
                    font-size:16px;
                    font-weight:700;
                    margin-top:5px;">
                    {project["Mandate"]}
                </div>

            </div>
            """)


        with c2:

            st.html(f"""
            <div style="
                background:white;
                border-radius:16px;
                padding:18px;
                border:1px solid #E5E7EB;
                box-shadow:0 5px 15px rgba(0,0,0,.05);
                font-family:Segoe UI,Arial,sans-serif;">

                <div style="
                    color:#6B7280;
                    font-size:12px;
                    font-weight:600;">
                    OWNER
                </div>

                <div style="
                    color:#006747;
                    font-size:16px;
                    font-weight:700;
                    margin-top:5px;">
                    {project["Allocation"]}
                </div>

            </div>
            """)


        with c3:

            st.html(f"""
            <div style="
                background:white;
                border-radius:16px;
                padding:18px;
                border:1px solid #E5E7EB;
                box-shadow:0 5px 15px rgba(0,0,0,.05);
                font-family:Segoe UI,Arial,sans-serif;">

                <div style="
                    color:#92400E;
                    font-size:12px;
                    font-weight:600;">
                    CURRENT STAGE
                </div>

                <div style="
                    color:#F59E0B;
                    font-size:16px;
                    font-weight:700;
                    margin-top:5px;">
                    {current}
                </div>

            </div>
            """)


        st.markdown("<br>", unsafe_allow_html=True)


    # ==========================================
    # DISPLAY PROJECTS
    # ==========================================

    if search_type == "Project":

        if selected == "All Projects":

            for _, project in timeline_df.iterrows():
                show_timeline(project)

        else:

            selected_project = timeline_df[
                timeline_df["Mandate"].astype(str) == selected
            ]

            if not selected_project.empty:
                show_timeline(selected_project.iloc[0])


    else:

        if selected == "All Members":

            for _, project in timeline_df.iterrows():
                show_timeline(project)

        else:

            member_df = timeline_df[
                timeline_df["Allocation"].astype(str) == selected
            ]

            st.success(
                f"{selected} is handling {len(member_df)} project(s)."
            )

            for _, project in member_df.iterrows():
                show_timeline(project)
# # =====================================================
# # TEAM PERFORMANCE
# # =====================================================

# elif page == "Team Performance":

#     # ==========================================
#     # HEADER
#     # ==========================================

#     st.html("""
#     <div style="
#         background:linear-gradient(180deg,#ffffff,#f8fbff);
#         border-radius:24px;
#         padding:28px;
#         border:1px solid #E5E7EB;
#         box-shadow:0 14px 35px rgba(0,0,0,.10);
#         margin-bottom:20px;
#         font-family:Segoe UI,Arial,sans-serif;">

#         <div style="
#             color:#006747;
#             font-size:14px;
#             font-weight:700;
#             letter-spacing:2px;
#             margin-bottom:8px;">
#             SMARTPAY PROJECT MANAGEMENT
#         </div>

#         <div style="
#             color:#006747;
#             font-size:40px;
#             font-weight:700;">
#             Team Performance
#         </div>

#         <div style="
#             color:#6B7280;
#             font-size:17px;
#             margin-top:10px;">
#             Monitor team workload, project distribution and delivery progress.
#         </div>

#     </div>
#     """)

#     st.markdown("<br>", unsafe_allow_html=True)


#     # ==========================================
#     # TEAM MEMBER FILTER
#     # ==========================================

#     st.markdown("""
#     <h2 style="
#         color:#006747;
#         font-size:28px;
#         margin-bottom:12px;">
#         👤 Team Member
#     </h2>
#     """, unsafe_allow_html=True)

#     member = st.selectbox(
#         "Select Team Member",
#         ["All"] + sorted(
#             df["Allocation"]
#             .dropna()
#             .astype(str)
#             .unique()
#         )
#     )

#     team_df = df.copy()

#     if member != "All":
#         team_df = team_df[
#             team_df["Allocation"].astype(str) == member
#         ]


#     # ==========================================
#     # CLEAN STATUS
#     # ==========================================

#     team_df["Status_Clean"] = (
#         team_df["Status"]
#         .astype(str)
#         .str.upper()
#         .str.strip()
#         .replace({
#             "SIT": "UNDER DEVELOPMENT",
#             "DEVELOPMENT": "UNDER DEVELOPMENT",
#             "SCOPING": "UNDER SCOPING",
#             "IS  REVIEW": "IS REVIEW"
#         })
#     )

#     status = team_df["Status_Clean"]


#     # ==========================================
#     # KPI VALUES
#     # ==========================================

#     total = len(team_df)
#     live = len(team_df[status == "LIVE"])
#     uat = len(team_df[status == "UAT"])
#     development = len(team_df[status == "UNDER DEVELOPMENT"])
#     review = len(team_df[status == "IS REVIEW"])
#     cmc = len(team_df[status == "CMC"])
#     scoping = len(team_df[status == "UNDER SCOPING"])


#     # ==========================
#     # KPI CARDS
#     # ==========================

#     k1, k2, k3, k4, k5, k6, k7 = st.columns(7)

#     with k1:
#         st.metric("Total", total)

#     with k2:
#         st.metric("Live", live)

#     with k3:
#         st.metric("UAT", uat)

#     with k4:
#         st.metric("Development", development)

#     with k5:
#         st.metric("IS Review", review)

#     with k6:
#         st.metric("CMC", cmc)

#     with k7:
#         st.metric("Scoping", scoping)

#     st.markdown("<br>", unsafe_allow_html=True)

#     # ==========================================
#     # PROJECT STATUS DISTRIBUTION
#     # ==========================================

#     st.markdown("""
#     <h2 style="
#         color:#006747;
#         font-size:30px;">
#         📊 Project Status Distribution
#     </h2>
#     """, unsafe_allow_html=True)


#     status_order = [
#         "LIVE",
#         "UAT",
#         "UNDER DEVELOPMENT",
#         "IS REVIEW",
#         "CMC",
#         "UNDER SCOPING"
#     ]


#     color_map = {
#         "LIVE": "#00C853",
#         "UAT": "#F9A825",
#         "UNDER DEVELOPMENT": "#FF9800",
#         "IS REVIEW": "#00ACC1",
#         "CMC": "#3949AB",
#         "UNDER SCOPING": "#8E24AA"
#     }


#     status_df = (
#         team_df["Status_Clean"]
#         .value_counts()
#         .reindex(status_order, fill_value=0)
#         .reset_index()
#     )

#     status_df.columns = [
#         "Status",
#         "Projects"
#     ]


#     fig1 = px.bar(
#         status_df,
#         x="Status",
#         y="Projects",
#         text="Projects",
#         color="Status",
#         color_discrete_map=color_map
#     )


#     # ==========================================
#     # CHART DESIGN + DARK MODE FIX
#     # ==========================================

#     fig1.update_layout(
#         height=500,

#         plot_bgcolor="white",
#         paper_bgcolor="white",

#         showlegend=False,

#         font=dict(
#             color="#111827",
#             family="Segoe UI, Arial"
#         ),

#         xaxis=dict(
#             title="",
#             tickangle=0,
#             tickfont=dict(
#                 size=11,
#                 color="#111827"
#             ),
#             automargin=True
#         ),

#         yaxis=dict(
#             title="Projects",
#             title_font=dict(
#                 color="#111827"
#             ),
#             tickfont=dict(
#                 color="#111827"
#             ),
#             gridcolor="#E5E7EB"
#         ),

#         margin=dict(
#             l=30,
#             r=30,
#             t=30,
#             b=100
#         )
#     )


#     fig1.update_traces(
#         textposition="outside",
#         textfont=dict(
#             color="#111827",
#             size=13
#         ),
#         marker_line_width=0
#     )


#     st.plotly_chart(
#         fig1,
#         width="stretch"
#     )


#     st.markdown("<br>", unsafe_allow_html=True)


#     # ==========================================
#     # TEAM WORKLOAD
#     # ==========================================

#     if member == "All":

#         st.markdown("""
#         <h2 style="
#             color:#006747;
#             font-size:30px;">
#             👥 Team Workload
#         </h2>
#         """, unsafe_allow_html=True)


#         allocation_df = (
#             df["Allocation"]
#             .value_counts()
#             .reset_index()
#         )

#         allocation_df.columns = [
#             "Allocation",
#             "Projects"
#         ]


#         fig2 = px.bar(
#             allocation_df,
#             x="Allocation",
#             y="Projects",
#             text="Projects",
#             color="Projects",
#             color_continuous_scale="Greens"
#         )


#         fig2.update_layout(
#             height=500,
#             plot_bgcolor="white",
#             paper_bgcolor="white",

#             font=dict(
#                 color="#111827",
#                 family="Segoe UI, Arial"
#             ),

#             coloraxis_showscale=False,

#             xaxis=dict(
#                 title="",
#                 tickfont=dict(
#                     color="#111827"
#                 )
#             ),

#             yaxis=dict(
#                 title="Projects",
#                 title_font=dict(
#                     color="#111827"
#                 ),
#                 tickfont=dict(
#                     color="#111827"
#                 ),
#                 gridcolor="#E5E7EB"
#             ),

#             margin=dict(
#                 l=20,
#                 r=20,
#                 t=25,
#                 b=30
#             )
#         )


#         fig2.update_traces(
#             textposition="outside",
#             textfont=dict(
#                 color="#111827",
#                 size=13
#             ),
#             marker_line_width=0
#         )


#         st.plotly_chart(
#             fig2,
#             width="stretch"
#         )


#     else:

#         # ==========================================
#         # SELECTED MEMBER PROJECTS - VIP
#         # ==========================================

#         st.markdown(f"""
#         <h2 style="
#         color:#006747;
#         font-size:30px;
#         font-weight:700;
#         margin-top:25px;
#         margin-bottom:15px;">
#         📋 {member} — Project Portfolio
#         </h2>
#         """, unsafe_allow_html=True)


#         # ==========================================
#         # PREPARE MEMBER DATA
#         # ==========================================

#         member_display_df = team_df[
#             [
#                 "Mandate",
#                 "Status",
#                 "Allocation"
#             ]
#         ].copy()


#         # ==========================================
#         # STATUS BADGE
#         # ==========================================

#         def member_status_badge(status):

#             status = str(status).strip()
#             status_upper = status.upper()

#             if status_upper == "LIVE":
#                 css = "status-live"

#             elif status_upper == "UAT":
#                 css = "status-uat"

#             elif status_upper in [
#                 "DEVELOPMENT",
#                 "UNDER DEVELOPMENT",
#                 "SIT"
#             ]:
#                 css = "status-development"

#             elif status_upper == "IS REVIEW":
#                 css = "status-review"

#             elif status_upper == "CMC":
#                 css = "status-cmc"

#             elif status_upper in [
#                 "SCOPING",
#                 "UNDER SCOPING"
#             ]:
#                 css = "status-scoping"

#             else:
#                 css = "status-default"

#             return f'<span class="status-badge {css}">{status}</span>'


#         # ==========================================
#         # APPLY VIP FORMATTING
#         # ==========================================

#         member_display_df["Status"] = (
#             member_display_df["Status"]
#             .apply(member_status_badge)
#         )


#         member_display_df["Mandate"] = (
#             member_display_df["Mandate"]
#             .apply(
#                 lambda x:
#                 f'<span class="project-name">📁 {x}</span>'
#             )
#         )


#         # ==========================================
#         # CREATE VIP HTML TABLE
#         # ==========================================

#         member_table_html = member_display_df.to_html(
#             index=False,
#             escape=False,
#             classes="project-table"
#         )


#         st.markdown(
#             f"""
#             <div class="project-table-wrapper">
#                 {member_table_html}
#             </div>
#             """,
#             unsafe_allow_html=True
#         )


#         st.markdown("<br>", unsafe_allow_html=True)


#     # ==========================================
#     # TOP WORKLOAD
#     # ==========================================

#     st.markdown("""
#     <h2 style="
#         color:#006747;
#         font-size:30px;">
#         🏆 Team Workload Leader
#     </h2>
#     """, unsafe_allow_html=True)


#     top = (
#         df["Allocation"]
#         .value_counts()
#         .reset_index()
#     )

#     top.columns = [
#         "Member",
#         "Projects"
#     ]


#     if not top.empty:

#         winner = top.iloc[0]


#         st.html(f"""
#         <div style="
#             background:linear-gradient(
#                 135deg,
#                 #ffffff,
#                 #f1f8f5
#             );
#             border-radius:22px;
#             padding:24px;
#             border:1px solid #D1FAE5;
#             border-left:7px solid #006747;
#             box-shadow:0 10px 28px rgba(0,0,0,.08);
#             font-family:Segoe UI,Arial,sans-serif;">

#             <div style="
#                 color:#6B7280;
#                 font-size:13px;
#                 font-weight:600;
#                 text-transform:uppercase;
#                 letter-spacing:1px;">
#                 Highest Project Workload
#             </div>

#             <div style="
#                 color:#006747;
#                 font-size:28px;
#                 font-weight:700;
#                 margin-top:8px;">
#                 🏆 {winner["Member"]}
#             </div>

#             <div style="
#                 color:#111827;
#                 font-size:17px;
#                 margin-top:5px;">
#                 Currently handling
#                 <b>{winner["Projects"]}</b>
#                 project(s)
#             </div>

#         </div>
#         """)
# # =====================================================
# # VOICE SEARCH
# # =====================================================

# elif page == "Voice Search":

#     # ==========================================
#     # VIP HEADER
#     # ==========================================

#     st.html("""
#     <div style="
#         background:linear-gradient(135deg,#ffffff,#f4fbf8);
#         border-radius:24px;
#         padding:30px;
#         border:1px solid #DDEBE5;
#         box-shadow:0 14px 35px rgba(0,0,0,.10);
#         margin-bottom:22px;
#         font-family:Segoe UI,Arial,sans-serif;">

#         <div style="
#             color:#006747;
#             font-size:14px;
#             font-weight:700;
#             letter-spacing:2px;
#             margin-bottom:8px;">
#             SMARTPAY INTELLIGENT SEARCH
#         </div>

#         <div style="
#             color:#006747;
#             font-size:40px;
#             font-weight:700;">
#             🎤 Voice Search
#         </div>

#         <div style="
#             color:#6B7280;
#             font-size:17px;
#             margin-top:10px;">
#             Find SmartPay projects instantly using your voice.
#         </div>

#     </div>
#     """)

#     # ==========================================
#     # HOW TO SEARCH
#     # ==========================================

#     st.html("""
#     <div style="
#         background:white;
#         border-radius:20px;
#         padding:22px;
#         border:1px solid #E5E7EB;
#         box-shadow:0 8px 22px rgba(0,0,0,.06);
#         margin-bottom:22px;">

#         <div style="
#             color:#006747;
#             font-size:18px;
#             font-weight:700;
#             margin-bottom:8px;">
#             🎙️ How to Search
#         </div>

#         <div style="
#             color:#4B5563;
#             font-size:15px;
#             line-height:1.7;">
#             Speak a <b>Project Name</b>, <b>Team Member</b>,
#             <b>Status</b>, <b>Category</b> or say
#             <b>"Show All Projects"</b>.
#         </div>

#     </div>
#     """)

#     # ==========================================
#     # VOICE COMMAND GUIDE
#     # ==========================================

#     st.markdown("""
#     <h2 style="
#         color:#006747;
#         font-size:28px;
#         margin-bottom:15px;">
#         💡 Voice Commands
#     </h2>
#     """, unsafe_allow_html=True)

#     vc1, vc2, vc3, vc4 = st.columns(4)

#     with vc1:

#         st.html("""
#         <div style="
#             background:#F0FDF4;
#             border:1px solid #BBF7D0;
#             border-radius:18px;
#             padding:20px;
#             min-height:125px;">

#             <div style="font-size:25px;">
#                 🟢
#             </div>

#             <div style="
#                 color:#166534;
#                 font-weight:700;
#                 margin-top:8px;">
#                 Live Projects
#             </div>

#             <div style="
#                 color:#6B7280;
#                 font-size:13px;
#                 margin-top:5px;">
#                 Say: <b>Live</b>
#             </div>

#         </div>
#         """)

#     with vc2:

#         st.html("""
#         <div style="
#             background:#FFFBEB;
#             border:1px solid #FDE68A;
#             border-radius:18px;
#             padding:20px;
#             min-height:125px;">

#             <div style="font-size:25px;">
#                 🟡
#             </div>

#             <div style="
#                 color:#92400E;
#                 font-weight:700;
#                 margin-top:8px;">
#                 UAT Projects
#             </div>

#             <div style="
#                 color:#6B7280;
#                 font-size:13px;
#                 margin-top:5px;">
#                 Say: <b>UAT</b>
#             </div>

#         </div>
#         """)

#     with vc3:

#         st.html("""
#         <div style="
#             background:#EFF6FF;
#             border:1px solid #BFDBFE;
#             border-radius:18px;
#             padding:20px;
#             min-height:125px;">

#             <div style="font-size:25px;">
#                 🔎
#             </div>

#             <div style="
#                 color:#1D4ED8;
#                 font-weight:700;
#                 margin-top:8px;">
#                 Project Search
#             </div>

#             <div style="
#                 color:#6B7280;
#                 font-size:13px;
#                 margin-top:5px;">
#                 Say the <b>project name</b>
#             </div>

#         </div>
#         """)

#     with vc4:

#         st.html("""
#         <div style="
#             background:#F5F3FF;
#             border:1px solid #DDD6FE;
#             border-radius:18px;
#             padding:20px;
#             min-height:125px;">

#             <div style="font-size:25px;">
#                 👥
#             </div>

#             <div style="
#                 color:#5B21B6;
#                 font-weight:700;
#                 margin-top:8px;">
#                 Team Search
#             </div>

#             <div style="
#                 color:#6B7280;
#                 font-size:13px;
#                 margin-top:5px;">
#                 Say the <b>team member</b>
#             </div>

#         </div>
#         """)

#     st.markdown("<br>", unsafe_allow_html=True)

#     # ==========================================
#     # MICROPHONE AREA
#     # ==========================================

#     st.html("""
#     <div style="
#         background:linear-gradient(135deg,#006747,#00875A);
#         border-radius:22px;
#         padding:25px;
#         text-align:center;
#         color:white;
#         box-shadow:0 12px 30px rgba(0,103,71,.20);
#         margin-bottom:20px;">

#         <div style="
#             font-size:42px;">
#             🎤
#         </div>

#         <div style="
#             font-size:22px;
#             font-weight:700;
#             margin-top:8px;">
#             Speak Your Command
#         </div>

#         <div style="
#             font-size:14px;
#             opacity:.9;
#             margin-top:6px;">
#             Use your microphone to search SmartPay projects
#         </div>

#     </div>
#     """)

#     # =====================================================
#     # ORIGINAL VOICE CODE — DO NOT CHANGE
#     # =====================================================

#     voice_text = listen()

#     st.write("Raw Voice :", repr(voice_text))

#     voice_text = normalize_voice(str(voice_text))

#     st.write("Normalized :", voice_text)

#     if voice_text:

#         voice_text = voice_text.lower().strip()

#         st.success(
#             f"🎤 You said: {voice_text}"
#         )

#         speak(
#             f"You said {voice_text}"
#         )

#         # =========================================
#         # SMART COMMANDS
#         # =========================================

#         if voice_text == "all":

#             result = df.copy()

#         elif voice_text == "live":

#             result = df[
#                 df["Status"]
#                 .astype(str)
#                 .str.lower() == "live"
#             ]

#         elif voice_text == "uat":

#             result = df[
#                 df["Status"]
#                 .astype(str)
#                 .str.lower() == "uat"
#             ]

#         elif voice_text == "sit":

#             result = df[
#                 df["Status"]
#                 .astype(str)
#                 .str.lower() == "sit"
#             ]

#         elif voice_text == "under development":

#             result = df[
#                 df["Status"]
#                 .astype(str)
#                 .str.lower() == "under development"
#             ]

#         else:

#             search_cols = [
#                 "Mandate",
#                 "Allocation",
#                 "Status",
#                 "Category",
#                 "Update"
#             ]

#             result = df[
#                 df.apply(
#                     lambda row: any(
#                         voice_text in str(
#                             row[col]
#                         ).lower()
#                         for col in search_cols
#                     ),
#                     axis=1
#                 )
#             ]

#             # =========================================
#             # FUZZY SEARCH
#             # =========================================

#             if result.empty:

#                 search_values = []

#                 for col in search_cols:

#                     search_values.extend(
#                         df[col]
#                         .dropna()
#                         .astype(str)
#                         .tolist()
#                     )

#                 match = process.extractOne(
#                     voice_text,
#                     search_values,
#                     scorer=fuzz.token_sort_ratio
#                 )

#                 if match and match[1] >= 70:

#                     matched = match[0].lower()

#                     result = df[
#                         df.apply(
#                             lambda row: any(
#                                 matched in str(
#                                     row[col]
#                                 ).lower()
#                                 for col in search_cols
#                             ),
#                             axis=1
#                         )
#                     ]

#         # =========================================
#         # RESULT
#         # =========================================

#         if result.empty:

#             st.error(
#                 "❌ No Project Found"
#             )

#             speak(
#                 "Sorry. No matching project found."
#             )

#         else:

#             st.success(
#                 f"✅ {len(result)} Project(s) Found"
#             )

#             st.metric(
#                 "Total Results",
#                 len(result)
#             )

#             # =========================================
#             # SINGLE RESULT
#             # =========================================

#             if len(result) == 1:

#                 first = result.iloc[0]

#                 status = str(
#                     first["Status"]
#                 ).upper()

#                 if status == "LIVE":

#                     badge = "#16A34A"

#                 elif status == "UAT":

#                     badge = "#F59E0B"

#                 elif status == "SIT":

#                     badge = "#2563EB"

#                 else:

#                     badge = "#6B7280"

#                 st.markdown(
#                     f"""
#                     <div style="
#                     background:white;
#                     border-radius:18px;
#                     padding:25px;
#                     border-left:8px solid {badge};
#                     box-shadow:0 8px 18px rgba(0,0,0,.08);
#                     margin-bottom:20px;">

#                     <h2 style="
#                     color:#006747;
#                     margin-top:0;">
#                     {first['Mandate']}
#                     </h2>

#                     <table style="
#                     width:100%;
#                     font-size:16px;">

#                     <tr>
#                     <td><b>Status</b></td>
#                     <td>{first['Status']}</td>
#                     </tr>

#                     <tr>
#                     <td><b>Owner</b></td>
#                     <td>{first['Allocation']}</td>
#                     </tr>

#                     <tr>
#                     <td><b>Category</b></td>
#                     <td>{first['Category']}</td>
#                     </tr>

#                     <tr>
#                     <td><b>Latest Update</b></td>
#                     <td>{first['Update']}</td>
#                     </tr>

#                     </table>

#                     </div>
#                     """,
#                     unsafe_allow_html=True
#                 )

#                 response = (
#                     f"{first['Mandate']} is currently "
#                     f"{first['Status']} and allocated to "
#                     f"{first['Allocation']}. "
#                     f"Latest update is "
#                     f"{first['Update']}."
#                 )

#             # =========================================
#             # MULTIPLE RESULTS
#             # =========================================

#             else:

#                 st.dataframe(
#                     result,
#                     use_container_width=True,
#                     hide_index=True
#                 )

#                 st.markdown(
#                     "### 📋 Projects Found"
#                 )

#                 c1, c2, c3 = st.columns(3)

#                 with c1:

#                     st.metric(
#                         "Projects",
#                         len(result)
#                     )

#                 with c2:

#                     st.metric(
#                         "Owners",
#                         result["Allocation"].nunique()
#                     )

#                 with c3:

#                     st.metric(
#                         "Live",
#                         len(
#                             result[
#                                 result["Status"]
#                                 .astype(str)
#                                 .str.upper() == "LIVE"
#                             ]
#                         )
#                     )

#                 for _, row in result.iterrows():

#                     st.markdown(
#                         f"""
#                         <div style="
#                         background:white;
#                         padding:18px;
#                         border-radius:15px;
#                         margin-bottom:12px;
#                         border:1px solid #E5E7EB;
#                         box-shadow:0 4px 10px rgba(0,0,0,.06);">

#                         <h4 style="
#                         color:#006747;
#                         margin:0;">
#                         {row['Mandate']}
#                         </h4>

#                         <p style="
#                         margin-top:8px;">

#                         <b>Owner:</b>
#                         {row['Allocation']}<br>

#                         <b>Status:</b>
#                         {row['Status']}<br>

#                         <b>Category:</b>
#                         {row['Category']}

#                         </p>

#                         </div>
#                         """,
#                         unsafe_allow_html=True
#                     )

#                 names = ", ".join(
#                     result["Mandate"]
#                     .astype(str)
#                     .tolist()
#                 )

#                 response = (
#                     f"{len(result)} projects found. "
#                     f"The projects are {names}."
#                 )

#             # =========================================
#             # VOICE RESPONSE
#             # =========================================

#             st.info(response)

#             speak(response)

# =====================================================
# BAU MONITORING
# =====================================================

elif page == "BAU Monitoring":

    # =================================================
    # HEADER
    # =================================================

    st.html("""
    <div style="
        background:linear-gradient(135deg,#ffffff,#f7fbf9);
        border-radius:24px;
        padding:30px;
        border:1px solid #E5E7EB;
        box-shadow:0 14px 35px rgba(0,0,0,.08);
        margin-bottom:20px;
        font-family:Segoe UI,Arial,sans-serif;
    ">

        <div style="
            color:#006747;
            font-size:14px;
            font-weight:700;
            letter-spacing:2px;
            margin-bottom:8px;">
            SMARTPAY BUSINESS AS USUAL
        </div>

        <div style="
            color:#006747;
            font-size:40px;
            font-weight:800;">
            🏦 BAU Monitoring
        </div>

        <div style="
            color:#6B7280;
            font-size:17px;
            margin-top:10px;">
            Live Business Operations Monitoring, ownership and ongoing updates.
        </div>

    </div>
    """)


    # =================================================
    # BAU DATA
    # =================================================

    bau_df = df[
        df["Status"]
        .astype(str)
        .str.strip()
        .str.upper()
        == "BAU"
    ].copy()


    # =================================================
    # KPI VALUES
    # =================================================

    total_bau = len(bau_df)

    bau_members = (
        bau_df["Allocation"]
        .dropna()
        .astype(str)
        .nunique()
    )

    bau_updates = (
        bau_df["Update"]
        .notna()
        .sum()
        if "Update" in bau_df.columns
        else 0
    )

    bau_categories = (
        bau_df["Category"]
        .dropna()
        .astype(str)
        .nunique()
        if "Category" in bau_df.columns
        else 0
    )


    # =================================================
    # KPI CARDS
    # =================================================

    k1, k2, k3, k4 = st.columns(4)


    def bau_card(title, value, color):

        st.html(
            f"""
            <div style="
                background:white;
                border-radius:20px;
                padding:22px 16px;
                height:135px;
                border-top:7px solid {color};
                box-shadow:0 8px 22px rgba(0,0,0,.08);
                display:flex;
                flex-direction:column;
                justify-content:space-between;
                align-items:center;
                text-align:center;
            ">

                <div style="
                    color:#6B7280;
                    font-size:14px;
                    font-weight:600;">
                    {title}
                </div>

                <div style="
                    color:{color};
                    font-size:44px;
                    font-weight:800;
                    line-height:1;">
                    {value}
                </div>

            </div>
            """
        )


    with k1:
        bau_card(
            "Total BAU Projects",
            total_bau,
            "#006747"
        )

    with k2:
        bau_card(
            "Team Members",
            bau_members,
            "#3949AB"
        )

    with k3:
        bau_card(
            "Updated Projects",
            bau_updates,
            "#F9A825"
        )

    with k4:
        bau_card(
            "Categories",
            bau_categories,
            "#00ACC1"
        )


    st.markdown("<br>", unsafe_allow_html=True)


    # =================================================
    # FILTERS
    # =================================================

    st.markdown("""
    <h2 style="
        color:#006747;
        font-size:28px;
        font-weight:700;
        margin-bottom:18px;">
        🎯 BAU Filters
    </h2>
    """, unsafe_allow_html=True)


    f1, f2 = st.columns(2)


    with f1:

        if "Allocation" in bau_df.columns:

            selected_bau_member = st.selectbox(
                "👤 Team Member",
                ["All"] +
                sorted(
                    bau_df["Allocation"]
                    .dropna()
                    .astype(str)
                    .unique()
                ),
                key="bau_member_filter"
            )

        else:

            selected_bau_member = "All"


    with f2:

        if "Category" in bau_df.columns:

            selected_bau_category = st.selectbox(
                "📂 Category",
                ["All"] +
                sorted(
                    bau_df["Category"]
                    .dropna()
                    .astype(str)
                    .unique()
                ),
                key="bau_category_filter"
            )

        else:

            selected_bau_category = "All"


    # =================================================
    # APPLY FILTERS
    # =================================================

    filtered_bau_df = bau_df.copy()


    if selected_bau_member != "All":

        filtered_bau_df = filtered_bau_df[
            filtered_bau_df["Allocation"].astype(str)
            == selected_bau_member
        ]


    if selected_bau_category != "All":

        filtered_bau_df = filtered_bau_df[
            filtered_bau_df["Category"].astype(str)
            == selected_bau_category
        ]


    # =================================================
    # SEARCH
    # =================================================

    bau_search = st.text_input(
        "",
        placeholder="🔍 Search BAU Project...",
        label_visibility="collapsed",
        key="bau_project_search"
    )


    if bau_search:

        filtered_bau_df = filtered_bau_df[
            filtered_bau_df["Mandate"]
            .astype(str)
            .str.contains(
                bau_search,
                case=False,
                na=False
            )
        ]


    st.markdown("<br>", unsafe_allow_html=True)


    # =================================================
    # OWNER SUMMARY
    # =================================================

    st.markdown("""
    <h2 style="
        color:#006747;
        font-size:28px;
        font-weight:700;
        margin-bottom:18px;">
        👥 BAU Owner Summary
    </h2>
    """, unsafe_allow_html=True)


    owner_summary = (
        filtered_bau_df
        .groupby("Allocation")
        .size()
        .reset_index(name="Projects")
        .sort_values(
            "Projects",
            ascending=False
        )
    )


    if owner_summary.empty:

        st.info("No BAU projects found.")

    else:

        owner_cols = st.columns(
            min(4, len(owner_summary))
        )


        for i, (_, owner) in enumerate(
            owner_summary.iterrows()
        ):

            with owner_cols[
                i % len(owner_cols)
            ]:

                st.html(
                    f"""
                    <div style="
                        background:linear-gradient(
                            180deg,
                            #ffffff,
                            #f7fbf9
                        );
                        border:1px solid #E5E7EB;
                        border-radius:20px;
                        padding:18px;
                        text-align:center;
                        box-shadow:
                            0 8px 20px rgba(0,0,0,.06);
                        margin-bottom:15px;
                    ">

                        <div style="
                            font-size:30px;">
                            👤
                        </div>

                        <div style="
                            color:#006747;
                            font-size:18px;
                            font-weight:700;
                            margin-top:6px;">
                            {owner["Allocation"]}
                        </div>

                        <div style="
                            color:#111827;
                            font-size:38px;
                            font-weight:800;
                            margin-top:6px;">
                            {owner["Projects"]}
                        </div>

                        <div style="
                            color:#6B7280;
                            font-size:13px;">
                            BAU Projects
                        </div>

                    </div>
                    """
                )


    # =================================================
    # BAU PROJECT TABLE
    # =================================================

    st.markdown("""
    <h2 style="
        color:#006747;
        font-size:30px;
        font-weight:700;
        margin-top:20px;
        margin-bottom:18px;">
        📋 BAU Project Portfolio
    </h2>
    """, unsafe_allow_html=True)


    display_bau = filtered_bau_df.copy()


    

    # =================================================
    # VIP TABLE CSS
    # =================================================

    st.markdown("""
    <style>

    .bau-table-wrapper {
        background:#FFFFFF;
        border:1px solid #DDE5E1;
        border-radius:18px;
        padding:6px;
        box-shadow:0 8px 25px rgba(0,103,71,.08);
        overflow-x:auto;
    }

    .bau-table {
        width:100%;
        border-collapse:separate;
        border-spacing:0;
        font-size:14px;
    }

    .bau-table thead th {
        background:#006747;
        color:white;
        padding:14px 12px;
        font-weight:700;
        text-align:left;
    }

    .bau-table tbody td {
        padding:13px 12px;
        color:#1F2937;
        border-bottom:1px solid #E5E7EB;
        background:#FFFFFF;
    }

    .bau-table tbody tr:nth-child(even) td {
        background:#F8FAFC;
    }

    .bau-table tbody tr:hover td {
        background:#ECFDF5;
    }

    .bau-project-name {
        color:#006747 !important;
        font-weight:700;
    }

    .bau-status {
        display:inline-block;
        padding:5px 12px;
        border-radius:20px;
        background:#E8F5E9;
        color:#166534;
        font-size:11px;
        font-weight:800;
    }

    </style>
    """, unsafe_allow_html=True)


    # =================================================
    # TABLE FORMATTING
    # =================================================

    if "Mandate" in display_bau.columns:

        display_bau["Mandate"] = (
            display_bau["Mandate"]
            .apply(
                lambda x:
                f'<span class="bau-project-name">📁 {x}</span>'
            )
        )


    if "Status" in display_bau.columns:

        display_bau["Status"] = (
            display_bau["Status"]
            .apply(
                lambda x:
                '<span class="bau-status">🟢 BAU</span>'
            )
        )


    # =================================================
    # SELECT USEFUL COLUMNS
    # =================================================

    preferred_columns = [
        "Mandate",
        "Allocation",
        "Status",
        "Category",
        "Update"
    ]


    table_columns = [
        col
        for col in preferred_columns
        if col in display_bau.columns
    ]


    if table_columns:

        display_bau = display_bau[
            table_columns
        ]


    # =================================================
    # CREATE TABLE
    # =================================================

    bau_table_html = display_bau.to_html(
        index=False,
        escape=False,
        classes="bau-table"
    )


    st.markdown(
        f"""
        <div class="bau-table-wrapper">
            {bau_table_html}
        </div>
        """,
        unsafe_allow_html=True
    )


    st.markdown("<br>", unsafe_allow_html=True)


    # =================================================
    # RECENT / ONGOING UPDATES
    # =================================================

    if "Update" in filtered_bau_df.columns:

        st.markdown("""
        <h2 style="
            color:#006747;
            font-size:28px;
            font-weight:700;
            margin-top:20px;
            margin-bottom:18px;">
            🔄 BAU Monitoring Updates
        </h2>
        """, unsafe_allow_html=True)


        updates_df = filtered_bau_df[
            [
                col
                for col in [
                    "Mandate",
                    "Allocation",
                    "Update"
                ]
                if col in filtered_bau_df.columns
            ]
        ].copy()


        if not updates_df.empty:

            for _, row in updates_df.iterrows():

                project_name = str(
                    row.get("Mandate", "")
                )

                owner_name = str(
                    row.get("Allocation", "")
                )

                update_text = str(
                    row.get(
                        "Update",
                        "Business as usual."
                    )
                )


                st.html(
                    f"""
                    <div style="
                        background:white;
                        border:1px solid #E5E7EB;
                        border-left:5px solid #006747;
                        border-radius:14px;
                        padding:15px 18px;
                        margin-bottom:10px;
                        box-shadow:0 4px 12px rgba(0,0,0,.05);
                    ">

                        <div style="
                            display:flex;
                            justify-content:space-between;
                            align-items:center;
                            gap:15px;
                        ">

                            <div style="
                                color:#006747;
                                font-size:16px;
                                font-weight:700;">
                                📁 {project_name}
                            </div>

                            <div style="
                                color:#6B7280;
                                font-size:12px;
                                font-weight:600;">
                                👤 {owner_name}
                            </div>

                        </div>

                        <div style="
                            color:#374151;
                            font-size:14px;
                            margin-top:9px;
                            line-height:1.5;">
                            {update_text}
                        </div>

                    </div>
                    """
                )

    else:

        st.info(
            "No BAU update field is available in the Excel data."
        )
# =====================================================
# EXPORT
# =====================================================

elif page == "Export":

    # ==========================================
    # HEADER
    # ==========================================

    st.html("""
    <div style="
        background:linear-gradient(135deg,#ffffff,#f4fbf8);
        border-radius:24px;
        padding:30px;
        border:1px solid #DDEBE5;
        box-shadow:0 14px 35px rgba(0,0,0,.10);
        margin-bottom:22px;
        font-family:Segoe UI,Arial,sans-serif;">

        <div style="
            color:#006747;
            font-size:14px;
            font-weight:700;
            letter-spacing:2px;
            margin-bottom:8px;">
            SMARTPAY PROJECT MANAGEMENT
        </div>

        <div style="
            color:#006747;
            font-size:40px;
            font-weight:700;">
            Executive Report Center
        </div>

        <div style="
            color:#6B7280;
            font-size:17px;
            margin-top:10px;">
            Generate, preview and download SmartPay project reports.
        </div>

    </div>
    """)

    st.markdown("<br>", unsafe_allow_html=True)


    # ==========================================
    # REPORT INFORMATION
    # ==========================================

    st.html("""
    <div style="
        background:white;
        border-radius:20px;
        padding:24px;
        border:1px solid #E5E7EB;
        box-shadow:0 8px 22px rgba(0,0,0,.06);
        margin-bottom:22px;">

        <div style="
            color:#006747;
            font-size:20px;
            font-weight:700;
            margin-bottom:15px;">
            📋 Report Includes
        </div>

        <div style="
            color:#374151;
            font-size:15px;
            line-height:2;">

            ✅ Dashboard Summary<br>
            ✅ KPI Overview<br>
            ✅ Project Details<br>
            ✅ Team Performance<br>
            ✅ Analytics Summary<br>
            ✅ Project Timeline

        </div>

    </div>
    """)


    # ==========================================
    # GENERATED DATE
    # ==========================================

    generated_time = datetime.now().strftime(
        "%d %B %Y  |  %I:%M %p"
    )

    st.html(f"""
    <div style="
        background:#F0FDF4;
        border:1px solid #BBF7D0;
        border-radius:16px;
        padding:16px 20px;
        margin-bottom:25px;">

        <div style="
            color:#166534;
            font-size:12px;
            font-weight:700;
            text-transform:uppercase;
            letter-spacing:1px;">
            Generated On
        </div>

        <div style="
            color:#166534;
            font-size:18px;
            font-weight:700;
            margin-top:5px;">
            {generated_time}
        </div>

    </div>
    """)


    # ==========================================
    # EXPORT OPTIONS
    # ==========================================

    st.markdown("""
    <h2 style="
        color:#006747;
        font-size:28px;
        margin-bottom:15px;">
        📤 Export Reports
    </h2>
    """, unsafe_allow_html=True)


    # ==========================================
    # CSV DATA
    # ==========================================

    csv = df.to_csv(
        index=False
    ).encode("utf-8")


    # ==========================================
    # EXCEL DATA
    # ==========================================

    excel_buffer = BytesIO()

    with pd.ExcelWriter(
        excel_buffer,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name="SmartPay Projects"
        )

        workbook = writer.book
        worksheet = writer.sheets[
            "SmartPay Projects"
        ]


        # ======================================
        # EXCEL TABLE
        # ======================================

        from openpyxl.worksheet.table import (
            Table,
            TableStyleInfo
        )

        from openpyxl.utils import (
            get_column_letter
        )


        last_row = worksheet.max_row
        last_col = worksheet.max_column

        last_col_letter = get_column_letter(
            last_col
        )

        table_ref = (
            f"A1:{last_col_letter}{last_row}"
        )


        tab = Table(
            displayName="SmartPayProjects",
            ref=table_ref
        )


        style = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )


        tab.tableStyleInfo = style

        worksheet.add_table(tab)


        # ======================================
        # FREEZE HEADER
        # ======================================

        worksheet.freeze_panes = "A2"


        # ======================================
        # AUTO COLUMN WIDTH
        # ======================================

        for column in worksheet.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                try:

                    if cell.value is not None:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except:

                    pass


            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                40
            )


    excel_data = (
        excel_buffer.getvalue()
    )


    # ==========================================
    # EXPORT CARDS
    # ==========================================

    c1, c2, c3 = st.columns(3)


    # ==========================================
    # CSV
    # ==========================================

    with c1:

        st.html("""
        <div style="
            background:#F8FAFC;
            border:1px solid #E5E7EB;
            border-top:5px solid #006747;
            border-radius:18px;
            padding:20px;
            min-height:105px;
            margin-bottom:10px;">

            <div style="font-size:26px;">
                📄
            </div>

            <div style="
                color:#006747;
                font-size:17px;
                font-weight:700;
                margin-top:7px;">
                CSV Report
            </div>

            <div style="
                color:#6B7280;
                font-size:13px;
                margin-top:4px;">
                Project data in CSV format
            </div>

        </div>
        """)


        st.download_button(
            "⬇ Export CSV",
            csv,
            "SmartPay_Projects.csv",
            "text/csv",
            use_container_width=True
        )


    # ==========================================
    # EXCEL
    # ==========================================

    with c2:

        st.html("""
        <div style="
            background:#F8FAFC;
            border:1px solid #E5E7EB;
            border-top:5px solid #16A34A;
            border-radius:18px;
            padding:20px;
            min-height:105px;
            margin-bottom:10px;">

            <div style="font-size:26px;">
                📊
            </div>

            <div style="
                color:#166534;
                font-size:17px;
                font-weight:700;
                margin-top:7px;">
                Excel Report
            </div>

            <div style="
                color:#6B7280;
                font-size:13px;
                margin-top:4px;">
                Formatted project data table
            </div>

        </div>
        """)


        st.download_button(
            "⬇ Export Excel",
            excel_data,
            "SmartPay_Projects.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


    # ==========================================
    # PDF
    # ==========================================

    with c3:

        st.html("""
        <div style="
            background:#F8FAFC;
            border:1px solid #E5E7EB;
            border-top:5px solid #DC2626;
            border-radius:18px;
            padding:20px;
            min-height:105px;
            margin-bottom:10px;">

            <div style="font-size:26px;">
                📑
            </div>

            <div style="
                color:#991B1B;
                font-size:17px;
                font-weight:700;
                margin-top:7px;">
                Executive PDF
            </div>

            <div style="
                color:#6B7280;
                font-size:13px;
                margin-top:4px;">
                Management-ready report
            </div>

        </div>
        """)


        if st.button(
            "📄 Generate Executive PDF",
            use_container_width=True
        ):

            pdf = generate_executive_pdf(df)


            st.download_button(
                "⬇ Download Executive PDF",
                data=pdf,
                file_name="SmartPay_Executive_Report.pdf",
                mime="application/pdf",
                use_container_width=True
            )


            st.success(
                "Executive PDF Generated Successfully."
            )


    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("---")


    # ==========================================
    # REPORT PREVIEW
    # ==========================================

    st.markdown("""
    <h2 style="
        color:#006747;
        font-size:28px;
        margin-bottom:15px;">
        👁️ Report Preview
    </h2>
    """, unsafe_allow_html=True)


    # ==========================================
    # SUMMARY CARD
    # ==========================================

    st.html(f"""
    <div style="
        background:white;
        border-radius:18px;
        padding:20px;
        border:1px solid #E5E7EB;
        box-shadow:0 6px 18px rgba(0,0,0,.06);
        margin-bottom:15px;">

        <div style="
            color:#6B7280;
            font-size:13px;
            font-weight:600;">
            TOTAL PROJECTS
        </div>

        <div style="
            color:#006747;
            font-size:34px;
            font-weight:700;
            margin-top:5px;">
            {len(df)}
        </div>

        <div style="
            color:#6B7280;
            font-size:14px;
            margin-top:5px;">
            Projects available for export
        </div>

    </div>
    """)


    # ==========================================
    # FULL DATA TABLE PREVIEW - VIP
    # ==========================================

    st.markdown("""
    <h3 style="
    color:#006747;
    font-size:24px;
    font-weight:700;
    margin-top:25px;
    margin-bottom:15px;">
    📊 Project Data Table
    </h3>
    """, unsafe_allow_html=True)


    # Copy data
    display_preview = df.copy()


    # ==========================================
    # STATUS BADGES
    # ==========================================

    def preview_status_badge(status):

        status = str(status).strip()
        status_upper = status.upper()

        if status_upper == "LIVE":
            css = "status-live"

        elif status_upper == "UAT":
            css = "status-uat"

        elif status_upper in [
            "DEVELOPMENT",
            "UNDER DEVELOPMENT",
            "SIT"
        ]:
            css = "status-development"

        elif status_upper == "IS REVIEW":
            css = "status-review"

        elif status_upper == "CMC":
            css = "status-cmc"

        elif status_upper in [
            "SCOPING",
            "UNDER SCOPING"
        ]:
            css = "status-scoping"

        else:
            css = "status-default"

        return f'<span class="status-badge {css}">{status}</span>'


    # Status
    if "Status" in display_preview.columns:

        display_preview["Status"] = (
            display_preview["Status"]
            .apply(preview_status_badge)
        )


    # Project name
    if "Mandate" in display_preview.columns:

        display_preview["Mandate"] = (
            display_preview["Mandate"]
            .apply(
                lambda x:
                f'<span class="project-name">📁 {x}</span>'
            )
        )


    # ==========================================
    # CREATE HTML TABLE
    # ==========================================

    preview_table_html = display_preview.to_html(
        index=False,
        escape=False,
        classes="project-table"
    )


    st.markdown(
        f"""
        <div class="project-table-wrapper">
            {preview_table_html}
        </div>
        """,
        unsafe_allow_html=True
    )
    # ==========================================
    # FINAL STATUS
    # ==========================================

    st.success(
        f"✅ Report Ready For Export — "
        f"{len(df)} Total Projects"
    )